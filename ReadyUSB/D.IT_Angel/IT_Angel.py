#!/usr/bin/env python3
"""
IT Angel - Automated Defensive IT Inspection & Monitoring
By SalgadoTech - IT-Tool Project
v3.0.0
"""

import os, sys, time, socket, datetime, platform, threading, subprocess, shutil

# --- Auto-install dependencies -------------------------------------------
def install_if_missing(packages):
    import importlib
    for pkg, import_name in packages:
        try:
            importlib.import_module(import_name)
        except ImportError:
            print(f"  [+] Installing {pkg}...")
            try:
                subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "--quiet"])
            except Exception as ex:
                print(f"  [!] Could not install {pkg}: {ex}")

install_if_missing([
    ("psutil",   "psutil"),
    ("openpyxl", "openpyxl"),
    ("colorama", "colorama"),
    ("rich",     "rich"),
    ("textual",  "textual"),
    ("pyserial", "serial"),
    ("pyusb",    "usb"),
])

import psutil, openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from colorama import init, Fore, Style
init(autoreset=True)
from rich.console import Console
from rich.table import Table
from rich.layout import Layout
from rich.panel import Panel
from rich.text import Text
from rich.live import Live
from rich import box
from textual.app import App, ComposeResult
from textual.containers import VerticalScroll
from textual.widgets import Static
from textual.binding import Binding

# --- Colors --------------------------------------------------------------
RED    = Fore.RED    + Style.BRIGHT
GREEN  = Fore.GREEN  + Style.BRIGHT
YELLOW = Fore.YELLOW + Style.BRIGHT
CYAN   = Fore.CYAN   + Style.BRIGHT
BLUE   = Fore.BLUE   + Style.BRIGHT
WHITE  = Fore.WHITE  + Style.BRIGHT
MAGENTA= Fore.MAGENTA+ Style.BRIGHT
RESET  = Style.RESET_ALL
DIM    = Style.DIM

# --- Excel styles --------------------------------------------------------
FILL_RED    = PatternFill("solid", fgColor="C0392B")
FILL_YELLOW = PatternFill("solid", fgColor="F39C12")
FILL_GREEN  = PatternFill("solid", fgColor="27AE60")
FILL_HEADER = PatternFill("solid", fgColor="0A0F1E")
FILL_SUBHDR = PatternFill("solid", fgColor="1A2C5B")
FILL_ALT    = PatternFill("solid", fgColor="EAF0FB")
FILL_WHITE  = PatternFill("solid", fgColor="FFFFFF")
FILL_PURPLE = PatternFill("solid", fgColor="6C3483")
FILL_TEAL   = PatternFill("solid", fgColor="117A65")

FONT_TITLE      = Font(name="Calibri", bold=True,  color="FFFFFF", size=14)
FONT_WHITE_BOLD = Font(name="Calibri", bold=True,  color="FFFFFF", size=11)
FONT_BLACK      = Font(name="Calibri",              color="000000", size=10)
FONT_WHITE      = Font(name="Calibri",              color="FFFFFF", size=10)
FONT_WHITE_B10  = Font(name="Calibri", bold=True,  color="FFFFFF", size=10)
FONT_BLACK_BOLD = Font(name="Calibri", bold=True,  color="000000", size=10)

thin_border = Border(
    left=Side(style='thin',   color="CCCCCC"),
    right=Side(style='thin',  color="CCCCCC"),
    top=Side(style='thin',    color="CCCCCC"),
    bottom=Side(style='thin', color="CCCCCC"),
)

# --- Globals -------------------------------------------------------------
VERSION        = "3.0.0"
g_os_target    = ""
g_duration     = 0        # seconds; 0 = indefinite
g_stop_event   = threading.Event()
g_excel_path   = ""
g_tshark_path  = ""       # resolved at startup

# Baselines
g_baseline_ports     = set()
g_baseline_processes = set()
g_baseline_services  = set()
g_baseline_usb       = set()
g_baseline_tasks     = set()

CYCLE_INTERVAL = 15 * 60  # 15 minutes

# --- Thread-safe Excel lock ----------------------------------------------
g_excel_lock = threading.Lock()

# --- BANNER --------------------------------------------------------------
def print_banner():
    os.system('cls' if os.name == 'nt' else 'clear 2>/dev/null')
    print(CYAN + r"   ___ _____   _   _   _  ____ _____ _     ")
    print(CYAN + r"  |_ _|_   _| / \ | \ | |/ ___| ____| |    ")
    print(CYAN + r"   | |  | |  / _ \|  \| | |  _|  _| | |    ")
    print(CYAN + r"   | |  | | / ___ \ |\  | |_| | |___| |___ ")
    print(CYAN + r"  |___| |_|/_/   \_\_| \_|\____|_____|_____|")
    print(CYAN + r"")
    print(WHITE  + f"  {'='*68}")
    print(CYAN   + f"  {'IT Angel  v' + VERSION:^68}")
    print(WHITE  + f"  {'Automated Defensive IT Inspection & Monitoring':^68}")
    print(DIM    + f"  {'By SalgadoTech  -  IT-Tool Project':^68}")
    print(WHITE  + f"  {'='*68}\n")


# --- TSHARK DETECTION ----------------------------------------------------
def detect_tshark():
    global g_tshark_path
    candidates_win = [
        r"C:\Program Files\Wireshark\tshark.exe",
        r"C:\Program Files (x86)\Wireshark\tshark.exe",
    ]
    if os.name == 'nt':
        for c in candidates_win:
            if os.path.isfile(c):
                g_tshark_path = c
                return True
        found = shutil.which("tshark")
        if found:
            g_tshark_path = found
            return True
    else:
        found = shutil.which("tshark")
        if found:
            g_tshark_path = found
            return True
    return False


def get_tshark_interface():
    """Find the tshark interface index whose IP matches the machine's active IP."""
    if not g_tshark_path:
        return "1"
    local_ip = get_local_ip()
    try:
        r = subprocess.run([g_tshark_path, "-D"],
                           capture_output=True, text=True, timeout=8)
        raw_lines = (r.stdout + r.stderr).strip().splitlines()
        iface_list = []
        for line in raw_lines:
            line = line.strip()
            if not line:
                continue
            parts = line.split(". ", 1)
            if len(parts) == 2 and parts[0].strip().isdigit():
                iface_list.append((parts[0].strip(), parts[1].strip()))
        if not iface_list:
            return "1"
        psutil_map = {}
        for iface_name, addrs in psutil.net_if_addrs().items():
            for a in addrs:
                if a.family == socket.AF_INET and a.address == local_ip:
                    psutil_map[iface_name.lower()] = a.address
        for idx, full_name in iface_list:
            friendly = full_name
            if "(" in full_name and full_name.endswith(")"):
                friendly = full_name[full_name.rfind("(")+1:-1]
            friendly_lower = friendly.lower()
            for ps_name in psutil_map:
                if ps_name in friendly_lower or friendly_lower in ps_name:
                    return idx
                ps_clean = ps_name.replace("-","").replace(" ","")
                fr_clean = friendly_lower.replace("-","").replace(" ","")
                if ps_clean in fr_clean or fr_clean in ps_clean:
                    return idx
        PREFER = ["wi-fi", "wifi", "wireless", "wlan", "ethernet",
                  "local area connection", "eth", "en", "eno", "ens", "enp"]
        AVOID  = ["loopback", "npcap loopback", "bluetooth", "vmware",
                  "virtualbox", "hyper-v", "any", "nflog", "nfqueue", "docker"]
        for idx, full_name in iface_list:
            name_l = full_name.lower()
            if any(a in name_l for a in AVOID):
                continue
            if any(p in name_l for p in PREFER):
                return idx
        for idx, full_name in iface_list:
            nl = full_name.lower()
            if "loopback" not in nl and nl not in ("any", "lo"):
                return idx
        return iface_list[0][0]
    except Exception:
        return "1"

# --- OS SELECTION --------------------------------------------------------
def select_os():
    global g_os_target
    print(WHITE + "  Which system will IT Angel protect today?\n")
    print(f"  {GREEN}[1]{RESET} Windows")
    print(f"  {GREEN}[2]{RESET} Linux\n")
    while True:
        c = input(f"  {CYAN}Select (1/2): {RESET}").strip()
        if c == "1":   g_os_target = "windows"; print(f"\n  {GREEN}Windows mode.{RESET}\n"); break
        elif c == "2": g_os_target = "linux";   print(f"\n  {GREEN}Linux mode.{RESET}\n");   break
        else:          print(f"  {RED}Enter 1 or 2.{RESET}")

# --- DURATION SELECTION --------------------------------------------------
def select_duration():
    global g_duration
    print(WHITE  + "  How many hours do you want IT Angel to protect this system?")
    print(YELLOW + "    - Type a number  (e.g.  2  or  0.5)")
    print(YELLOW + "    - Type  F        for a quick Fast Check")
    print(YELLOW + "    - Type  I        for Indefinite protection\n")
    while True:
        raw = input(f"  {CYAN}Your choice: {RESET}").strip().upper()
        if raw == "F":
            g_duration = -1
            print(f"\n  {MAGENTA}Fast Check mode selected.{RESET}\n")
            break
        elif raw == "I":
            g_duration = 0
            print(f"\n  {GREEN}Indefinite protection. Press Ctrl+C anytime to stop.{RESET}\n")
            break
        else:
            try:
                h = float(raw)
                if h <= 0: raise ValueError
                g_duration = int(h * 3600)
                label = f"{h:.1f}".rstrip('0').rstrip('.') + " hour(s)"
                print(f"\n  {GREEN}IT Angel will protect this system for {label}.{RESET}\n")
                break
            except ValueError:
                print(f"  {RED}Invalid input. Enter a number, F, or I.{RESET}")

# --- HELPERS -------------------------------------------------------------
def now_str():
    return datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def get_loginctl_sessions():
    """Linux: enumerate active user sessions via loginctl (works with systemd
    /xrdp/ssh where utmp - who / psutil.users() - is often empty). Returns a
    list of dicts: {user, session, remote(bool), host, type, tty, seat}."""
    sessions = []
    ids = run_cmd("loginctl list-sessions --no-legend 2>/dev/null", timeout=6)
    if not ids:
        return sessions
    for line in ids.splitlines():
        parts = line.split()
        if not parts:
            continue
        sid = parts[0]
        det = run_cmd(f"loginctl show-session {sid} "
                      f"-p Name -p Remote -p RemoteHost -p Type -p TTY -p Seat "
                      f"-p State -p Class 2>/dev/null", timeout=6)
        d = {}
        for l in det.splitlines():
            if "=" in l:
                k, v = l.split("=", 1)
                d[k.strip()] = v.strip()
        if d.get("Class") and d.get("Class") != "user":
            continue
        sessions.append({
            "user":    d.get("Name", "?"),
            "session": sid,
            "remote":  d.get("Remote", "no") == "yes",
            "host":    d.get("RemoteHost", "") or "",
            "type":    d.get("Type", ""),
            "tty":     d.get("TTY", "") or d.get("Seat", ""),
            "state":   d.get("State", ""),
        })
    return sessions

def run_cmd(cmd, shell=True, timeout=15):
    try:
        r = subprocess.run(cmd, shell=shell, capture_output=True, text=True, timeout=timeout)
        return r.stdout.strip()
    except Exception:
        return ""

def get_desktop_path():
    if os.name == 'nt':
        return os.path.join(os.path.expanduser("~"), "Desktop")
    sudo_user = os.environ.get("SUDO_USER")
    if sudo_user:
        for candidate in [f"/home/{sudo_user}/Desktop", f"/home/{sudo_user}"]:
            if os.path.exists(candidate):
                return candidate
    d = os.path.join(os.path.expanduser("~"), "Desktop")
    return d if os.path.exists(d) else os.path.expanduser("~")


# --- EXCEL SETUP ---------------------------------------------------------
def setup_excel():
    global g_excel_path
    ts       = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    hostname = socket.gethostname()
    fname    = f"IT_Angel_{hostname}_{ts}.xlsx"
    g_excel_path = os.path.join(get_desktop_path(), fname)

    wb      = openpyxl.Workbook()
    ws_tech = wb.active
    ws_tech.title = "Technical Log"
    ws_tech.sheet_view.showGridLines = False

    ws_tech.merge_cells("A1:J1")
    c = ws_tech["A1"]
    c.value     = f"IT ANGEL - TECHNICAL LOG  |  Host: {hostname}  |  OS: {g_os_target.upper()}  |  Started: {now_str()}"
    c.fill      = FILL_HEADER
    c.font      = FONT_TITLE
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_tech.row_dimensions[1].height = 30

    headers = ["Timestamp","Cycle","Category","Check","Finding","Details","Status","Delta","Severity","Recommendation"]
    for col, h in enumerate(headers, 1):
        cell = ws_tech.cell(row=2, column=col, value=h)
        cell.fill      = FILL_SUBHDR
        cell.font      = FONT_WHITE_BOLD
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border
    ws_tech.row_dimensions[2].height = 22
    for i, w in enumerate([20,8,18,28,35,45,12,10,12,35], 1):
        ws_tech.column_dimensions[get_column_letter(i)].width = w
    ws_tech.freeze_panes = "A3"

    ws_exec = wb.create_sheet("Executive Summary")
    ws_exec.sheet_view.showGridLines = False
    ws_exec.merge_cells("A1:F1")
    e = ws_exec["A1"]
    e.value     = f"IT ANGEL - EXECUTIVE SUMMARY  |  {hostname}  |  {g_os_target.upper()}"
    e.fill      = FILL_HEADER
    e.font      = FONT_TITLE
    e.alignment = Alignment(horizontal="center", vertical="center")
    ws_exec.row_dimensions[1].height = 30
    for col, h in enumerate(["Timestamp","Cycle","Category","Status","Critical Findings","Recommendation"], 1):
        cell = ws_exec.cell(row=2, column=col, value=h)
        cell.fill      = FILL_SUBHDR
        cell.font      = FONT_WHITE_BOLD
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = thin_border
    ws_exec.row_dimensions[2].height = 22
    for i, w in enumerate([20,8,22,14,50,40], 1):
        ws_exec.column_dimensions[get_column_letter(i)].width = w
    ws_exec.freeze_panes = "A3"

    ws_net = wb.create_sheet("Network Traffic")
    ws_net.sheet_view.showGridLines = False
    ws_net.merge_cells("A1:F1")
    n = ws_net["A1"]
    n.value     = f"IT ANGEL - NETWORK TRAFFIC LOG  |  {hostname}  |  {g_os_target.upper()}"
    n.fill      = FILL_PURPLE
    n.font      = FONT_TITLE
    n.alignment = Alignment(horizontal="center", vertical="center")
    ws_net.row_dimensions[1].height = 30
    for col, h in enumerate(["Timestamp","Cycle","Source","Destination","Protocol","Info"], 1):
        cell = ws_net.cell(row=2, column=col, value=h)
        cell.fill      = FILL_TEAL
        cell.font      = FONT_WHITE_BOLD
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = thin_border
    ws_net.row_dimensions[2].height = 22
    for i, w in enumerate([20,8,22,22,14,50], 1):
        ws_net.column_dimensions[get_column_letter(i)].width = w
    ws_net.freeze_panes = "A3"

    wb.save(g_excel_path)
    print(f"  {GREEN}Report file created:{RESET} {g_excel_path}\n")

# --- EXCEL APPEND --------------------------------------------------------
def append_to_excel(tech_rows, exec_rows, traffic_rows=None):
    with g_excel_lock:
        try:
            wb      = openpyxl.load_workbook(g_excel_path)
            ws_tech = wb["Technical Log"]
            ws_exec = wb["Executive Summary"]
            ws_net  = wb["Network Traffic"]

            next_r = ws_tech.max_row + 1
            for i, row in enumerate(tech_rows):
                r   = next_r + i
                alt = (r % 2 == 0)
                sev = row.get("Severity","").upper()
                vals = [row.get("Timestamp",""), row.get("Cycle",""), row.get("Category",""),
                        row.get("Check",""),     row.get("Finding",""), row.get("Details",""),
                        row.get("Status",""),    row.get("Delta",""),   row.get("Severity",""),
                        row.get("Recommendation","")]
                for col, val in enumerate(vals, 1):
                    cell = ws_tech.cell(row=r, column=col, value=val)
                    cell.border    = thin_border
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    if sev == "CRITICAL":
                        cell.fill = FILL_RED;    cell.font = FONT_WHITE_B10
                    elif sev == "WARNING":
                        cell.fill = FILL_YELLOW; cell.font = FONT_BLACK_BOLD
                    elif sev == "INFO":
                        cell.fill = FILL_ALT;    cell.font = FONT_BLACK
                    else:
                        cell.fill = FILL_ALT if alt else FILL_WHITE
                        cell.font = FONT_BLACK
                ws_tech.row_dimensions[r].height = 18

            next_e = ws_exec.max_row + 1
            for i, row in enumerate(exec_rows):
                r   = next_e + i
                sev = row.get("Status","").upper()
                vals = [row.get("Timestamp",""), row.get("Cycle",""), row.get("Category",""),
                        row.get("Status",""),    row.get("Critical Findings",""), row.get("Recommendation","")]
                for col, val in enumerate(vals, 1):
                    cell = ws_exec.cell(row=r, column=col, value=val)
                    cell.border    = thin_border
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    if sev == "CRITICAL":
                        cell.fill = FILL_RED;    cell.font = FONT_WHITE_B10
                    elif sev == "WARNING":
                        cell.fill = FILL_YELLOW; cell.font = FONT_BLACK_BOLD
                    elif sev == "OK":
                        cell.fill = FILL_GREEN;  cell.font = FONT_WHITE
                    else:
                        cell.fill = FILL_WHITE;  cell.font = FONT_BLACK
                ws_exec.row_dimensions[r].height = 18

            if traffic_rows:
                next_n = ws_net.max_row + 1
                for i, row in enumerate(traffic_rows):
                    r = next_n + i
                    alt = (r % 2 == 0)
                    vals = [row.get("Timestamp",""), row.get("Cycle",""), row.get("Source",""),
                            row.get("Destination",""), row.get("Protocol",""), row.get("Info","")]
                    for col, val in enumerate(vals, 1):
                        cell = ws_net.cell(row=r, column=col, value=val)
                        cell.border    = thin_border
                        cell.alignment = Alignment(vertical="center", wrap_text=True)
                        cell.fill      = FILL_ALT if alt else FILL_WHITE
                        cell.font      = FONT_BLACK
                    ws_net.row_dimensions[r].height = 16

            wb.save(g_excel_path)
        except Exception as ex:
            print(f"  {RED}[Excel Error] {ex}{RESET}")

# =========================================================================
#  CHECK MODULES
# =========================================================================

def check_system_snapshot():
    ts = now_str(); rows_t = []; rows_e = []
    hostname = socket.gethostname()
    ip       = get_local_ip()
    uname      = platform.uname()
    uptime_sec = time.time() - psutil.boot_time()
    uptime_str = str(datetime.timedelta(seconds=int(uptime_sec)))
    user       = os.environ.get("USERNAME") or os.environ.get("USER") or "N/A"
    ram_gb     = round(psutil.virtual_memory().total / (1024**3), 2)
    for check, finding in [
        ("Hostname",    hostname),
        ("IP Address",  ip),
        ("OS",          f"{uname.system} {uname.release} {uname.machine}"),
        ("Current User",user),
        ("Uptime",      uptime_str),
        ("CPU Cores",   str(psutil.cpu_count(logical=True))),
        ("RAM",         f"{ram_gb} GB"),
    ]:
        rows_t.append({"Timestamp":ts,"Cycle":"INIT","Category":"System Snapshot",
                        "Check":check,"Finding":finding,"Details":"",
                        "Status":"OK","Delta":"BASELINE","Severity":"INFO","Recommendation":""})
    rows_e.append({"Timestamp":ts,"Cycle":"INIT","Category":"System Snapshot","Status":"OK",
                   "Critical Findings":f"Host:{hostname} | IP:{ip} | User:{user} | Uptime:{uptime_str}",
                   "Recommendation":"Baseline captured."})
    return rows_t, rows_e


def check_network_adapters():
    ts = now_str(); rows_t = []; rows_e = []
    addrs = psutil.net_if_addrs(); stats = psutil.net_if_stats()
    for iface, addr_list in addrs.items():
        for addr in addr_list:
            if addr.family == socket.AF_INET:
                up  = stats[iface].isup if iface in stats else False
                sev = "OK" if up else "WARNING"
                rows_t.append({"Timestamp":ts,"Cycle":"INIT","Category":"Network Adapters",
                                "Check":iface,"Finding":addr.address,
                                "Details":f"Netmask:{addr.netmask} | Up:{up}",
                                "Status":"UP" if up else "DOWN","Delta":"BASELINE",
                                "Severity":sev,
                                "Recommendation":"" if up else f"Interface {iface} is DOWN"})
    findings = ", ".join([f"{i}:{a.address}" for i,al in addrs.items() for a in al if a.family==socket.AF_INET])
    rows_e.append({"Timestamp":ts,"Cycle":"INIT","Category":"Network Adapters","Status":"OK",
                   "Critical Findings":findings,"Recommendation":"Adapters catalogued."})
    return rows_t, rows_e


def check_active_connections(cycle):
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    SUSPICIOUS = {4444,1337,31337,6666,9999,5555,8888,1234,65535,31338,12345,54321}
    total = 0
    try:
        seen = set()
        for c in psutil.net_connections(kind='inet'):
            if c.status not in ('ESTABLISHED','LISTEN'): continue
            try:    pname = psutil.Process(c.pid).name()
            except: pname = "N/A"
            laddr = f"{c.laddr.ip}:{c.laddr.port}" if c.laddr else ""
            raddr = f"{c.raddr.ip}:{c.raddr.port}" if c.raddr else "-"
            key   = (laddr, raddr, pname)
            if key in seen: continue
            seen.add(key); total += 1
            port = c.laddr.port if c.laddr else 0
            sev  = "OK"; rec = ""
            if port in SUSPICIOUS or (c.raddr and c.raddr.port in SUSPICIOUS):
                sev = "CRITICAL"
                rec = f"Suspicious port {port} - verify {pname}"
                critical_findings.append(f"SUSPICIOUS PORT {port} by {pname}")
                rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Active Connections",
                                "Check":f"{c.status} {laddr}","Finding":f"-> {raddr}",
                                "Details":f"Process:{pname} PID:{c.pid}",
                                "Status":c.status,"Delta":"","Severity":sev,"Recommendation":rec})
    except Exception as ex:
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Active Connections",
                        "Check":"Error","Finding":str(ex),"Details":"",
                        "Status":"ERROR","Delta":"","Severity":"WARNING",
                        "Recommendation":"Run as administrator/root"})
    rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Active Connections",
                    "Check":"Summary","Finding":f"{total} connections active",
                    "Details":"Only suspicious connections logged above",
                    "Status":"OK","Delta":"","Severity":"OK","Recommendation":""})
    status = "CRITICAL" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"Active Connections","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else f"{total} connections - none suspicious",
                   "Recommendation":"Investigate flagged ports" if critical_findings else "Normal"})
    return rows_t, rows_e


def check_listening_ports(cycle):
    global g_baseline_ports
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    HIGH_RISK = {4444,1337,31337,6666,9999,5555,8888,23,512,513,514}
    current = set()
    try:
        for c in psutil.net_connections(kind='inet'):
            if c.status == 'LISTEN' and c.laddr:
                current.add(c.laddr.port)
    except: pass
    is_baseline = len(g_baseline_ports) == 0
    new_ports   = current - g_baseline_ports
    gone_ports  = g_baseline_ports - current
    for p in current:
        if p in HIGH_RISK:
            critical_findings.append(f"HIGH-RISK PORT: {p}")
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Listening Ports",
                            "Check":f"Port {p}","Finding":"HIGH-RISK OPEN","Details":"",
                            "Status":"CRITICAL","Delta":"","Severity":"CRITICAL",
                            "Recommendation":f"High-risk port {p} - verify immediately"})
    if is_baseline:
        g_baseline_ports = current
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Listening Ports",
                        "Check":"Baseline","Finding":f"{len(current)} ports catalogued",
                        "Details":f"Ports: {sorted(current)}",
                        "Status":"OK","Delta":"BASELINE","Severity":"INFO","Recommendation":""})
    else:
        for p in new_ports:
            sev = "CRITICAL" if p in HIGH_RISK else "WARNING"
            if p not in HIGH_RISK:
                critical_findings.append(f"NEW PORT: {p}")
                rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Listening Ports",
                                "Check":f"Port {p}","Finding":"NEW - appeared this cycle",
                                "Details":"","Status":"LISTEN","Delta":"NEW",
                                "Severity":sev,"Recommendation":f"New listening port {p} - verify"})
        for p in gone_ports:
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Listening Ports",
                            "Check":f"Port {p}","Finding":"CLOSED","Details":"Was open last cycle",
                            "Status":"CLOSED","Delta":"GONE","Severity":"INFO",
                            "Recommendation":"Verify if expected"})
        if not new_ports and not gone_ports and not critical_findings:
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Listening Ports",
                            "Check":"Status","Finding":f"{len(current)} ports - no changes",
                            "Details":"","Status":"OK","Delta":"","Severity":"OK","Recommendation":""})
        g_baseline_ports = current
    status = "CRITICAL" if any(p in HIGH_RISK for p in current) else "WARNING" if new_ports else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"Listening Ports","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else f"{len(current)} ports - no anomalies",
                   "Recommendation":"Investigate" if critical_findings else "Normal"})
    return rows_t, rows_e

def check_network_processes(cycle):
    global g_baseline_processes
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    SUSPICIOUS_EXACT = {
        "nc.exe","nc","ncat","ncat.exe","netcat","netcat.exe",
        "meterpreter","mimikatz","mimikatz.exe","empire","cobalt",
        "metasploit","nmap","nmap.exe","masscan","masscan.exe",
        "psexec","psexec.exe","wce","wce.exe","pwdump","fgdump",
        "gsecdump","quarks-pwdump","procdump.exe"
    }
    is_baseline = len(g_baseline_processes) == 0
    current = set()
    try:
        pids = {c.pid for c in psutil.net_connections(kind='inet') if c.pid}
        for pid in pids:
            try:
                p     = psutil.Process(pid)
                pname = p.name()
                current.add(pname)
                is_suspicious = pname.lower() in SUSPICIOUS_EXACT
                is_new = pname not in g_baseline_processes and not is_baseline
                if is_suspicious:
                    rec = f"{pname} is a known attack tool - investigate immediately"
                    critical_findings.append(f"ATTACK TOOL: {pname} PID:{pid}")
                    try:    cmd = " ".join(p.cmdline()[:4])
                    except: cmd = "N/A"
                    rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Net Processes",
                                    "Check":pname,"Finding":f"PID {pid}","Details":f"CMD:{cmd}",
                                    "Status":"SUSPICIOUS","Delta":"","Severity":"CRITICAL","Recommendation":rec})
                elif is_new:
                    rec = f"{pname} newly appeared with network activity"
                    critical_findings.append(f"NEW NET PROCESS: {pname}")
                    try:    cmd = " ".join(p.cmdline()[:4])
                    except: cmd = "N/A"
                    rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Net Processes",
                                    "Check":pname,"Finding":f"PID {pid}","Details":f"CMD:{cmd}",
                                    "Status":"NEW","Delta":"NEW","Severity":"WARNING","Recommendation":rec})
            except: pass
    except: pass
    if is_baseline:
        g_baseline_processes = current
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Net Processes",
                        "Check":"Baseline","Finding":f"{len(current)} processes with network activity",
                        "Details":"New processes will be flagged in next cycles",
                        "Status":"OK","Delta":"BASELINE","Severity":"INFO","Recommendation":""})
    else:
        g_baseline_processes = current
        if not critical_findings:
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Net Processes",
                            "Check":"Status","Finding":f"{len(current)} processes - no new or suspicious",
                            "Details":"","Status":"OK","Delta":"","Severity":"OK","Recommendation":""})
    status = "CRITICAL" if any("ATTACK" in f for f in critical_findings) else \
             "WARNING" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"Net Processes","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else f"{len(current)} processes - no threats",
                   "Recommendation":"Investigate" if critical_findings else "Normal"})
    return rows_t, rows_e


def check_resources(cycle):
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    cpu  = psutil.cpu_percent(interval=1)
    ram  = psutil.virtual_memory()
    disk = psutil.disk_usage('/')
    for check, finding, is_crit, is_warn, rec in [
        ("CPU Usage",     f"{cpu}%",              cpu>90,         cpu>75,         "CPU >90% - check runaway processes"),
        ("RAM Usage",     f"{ram.percent}%",       ram.percent>90, ram.percent>75, "RAM >90% - investigate memory hogs"),
        ("Disk Usage",    f"{disk.percent}%",      disk.percent>95,disk.percent>85,"Disk almost full - clean up"),
        ("RAM Available", f"{round(ram.available/1024**3,2)} GB", False, False,   ""),
        ("Disk Free",     f"{round(disk.free/1024**3,2)} GB",     False, False,   ""),
    ]:
        sev = "CRITICAL" if is_crit else "WARNING" if is_warn else "OK"
        if is_crit: critical_findings.append(f"{check}:{finding}")
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Resources",
                        "Check":check,"Finding":finding,"Details":"",
                        "Status":sev,"Delta":"","Severity":sev,
                        "Recommendation":rec if (is_crit or is_warn) else ""})
    try:
        top = sorted(psutil.process_iter(['name','cpu_percent']), key=lambda p: p.info['cpu_percent'] or 0, reverse=True)[:5]
        for p in top:
            if (p.info['cpu_percent'] or 0) > 0:
                rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Resources",
                                "Check":f"Top CPU: {p.info['name']}",
                                "Finding":f"{p.info['cpu_percent']}%","Details":"",
                                "Status":"INFO","Delta":"","Severity":"INFO","Recommendation":""})
    except: pass
    status = "CRITICAL" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"Resources","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else f"CPU:{cpu}% RAM:{ram.percent}% Disk:{disk.percent}%",
                   "Recommendation":"Investigate" if critical_findings else "Normal"})
    return rows_t, rows_e


def check_logged_users(cycle):
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    count = 0
    users = psutil.users()
    for u in users:
        login_t    = datetime.datetime.fromtimestamp(u.started).strftime("%Y-%m-%d %H:%M:%S")
        suspicious = u.host not in ('','::1','localhost','127.0.0.1',':0') and bool(u.host)
        sev        = "WARNING" if suspicious else "OK"
        if suspicious: critical_findings.append(f"REMOTE LOGIN: {u.name} from {u.host}")
        src = u.host if (u.host and u.host not in ('','::1','localhost','127.0.0.1',':0')) else "local"
        count += 1
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Logged Users",
                        "Check":u.name,"Finding":src,"Details":f"Login:{login_t}",
                        "Status":"REMOTE" if suspicious else "LOCAL","Delta":"","Severity":sev,
                        "Recommendation":f"Verify remote session from {u.host}" if suspicious else ""})
    # Linux fallback: systemd / xrdp / ssh sessions are frequently absent from
    # utmp, so psutil.users() returns []. Enumerate them via loginctl instead.
    if not users and g_os_target != "windows":
        for s in get_loginctl_sessions():
            suspicious = bool(s["remote"]) and s["host"] not in ('','::1','localhost','127.0.0.1')
            sev = "WARNING" if suspicious else "OK"
            if suspicious: critical_findings.append(f"REMOTE LOGIN: {s['user']} from {s['host']}")
            src = s["host"] if suspicious else "local"
            count += 1
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Logged Users",
                            "Check":s["user"],"Finding":src,
                            "Details":f"Session:{s['session']} Type:{s['type']} TTY:{s['tty']} State:{s['state']}",
                            "Status":"REMOTE" if suspicious else "LOCAL","Delta":"","Severity":sev,
                            "Recommendation":f"Verify remote session from {s['host']}" if suspicious else ""})
    if count == 0:
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Logged Users",
                        "Check":"No users","Finding":"N/A","Details":"",
                        "Status":"OK","Delta":"","Severity":"INFO","Recommendation":""})
    status = "WARNING" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"Logged Users","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else f"{count} user(s) active",
                   "Recommendation":"Verify remote sessions" if critical_findings else "Normal"})
    return rows_t, rows_e

def check_services(cycle):
    global g_baseline_services
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []; current = set()
    CRITICAL_WIN = {"WinDefend", "MpsSvc", "EventLog", "Spooler", "LanmanServer"}
    CRITICAL_LIN = {"ssh","ufw","firewalld","cron","rsyslog","auditd","fail2ban","networking"}
    is_baseline  = len(g_baseline_services) == 0
    if g_os_target == "windows":
        try:
            for svc in psutil.win_service_iter():
                try:
                    info       = svc.as_dict()
                    name       = info['name']
                    status_svc = info['status']
                    current.add(name)
                    if is_baseline:
                        if name in CRITICAL_WIN and status_svc != "running":
                            critical_findings.append(f"SERVICE DOWN: {name}")
                            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Services",
                                            "Check":name,"Finding":status_svc,"Details":info.get('display_name',''),
                                            "Status":status_svc.upper(),"Delta":"BASELINE","Severity":"CRITICAL",
                                            "Recommendation":f"Critical service {name} is {status_svc}"})
                    else:
                        is_new = name not in g_baseline_services
                        if name in CRITICAL_WIN and status_svc != "running":
                            critical_findings.append(f"SERVICE DOWN: {name}")
                            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Services",
                                            "Check":name,"Finding":status_svc,"Details":info.get('display_name',''),
                                            "Status":status_svc.upper(),"Delta":"","Severity":"CRITICAL",
                                            "Recommendation":f"Critical service {name} is {status_svc}"})
                        elif is_new:
                            critical_findings.append(f"NEW SERVICE: {name}")
                            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Services",
                                            "Check":name,"Finding":status_svc,"Details":info.get('display_name',''),
                                            "Status":status_svc.upper(),"Delta":"NEW","Severity":"WARNING",
                                            "Recommendation":f"New service {name} - verify"})
                except: pass
            if is_baseline:
                rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Services",
                                "Check":"Baseline","Finding":f"{len(current)} services catalogued",
                                "Details":"Only critical failures and NEW services reported going forward",
                                "Status":"OK","Delta":"BASELINE","Severity":"INFO","Recommendation":""})
        except Exception as ex:
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Services",
                            "Check":"Error","Finding":str(ex),"Details":"",
                            "Status":"ERROR","Delta":"","Severity":"WARNING","Recommendation":"Run as administrator"})
    else:
        out = run_cmd("systemctl list-units --type=service --state=running --no-pager --plain 2>/dev/null | awk '{print $1}'")
        for line in out.splitlines():
            name = line.strip().replace(".service","")
            if not name: continue
            current.add(name)
            if not is_baseline:
                is_new = name not in g_baseline_services
                if is_new:
                    critical_findings.append(f"NEW SERVICE: {name}")
                    rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Services",
                                    "Check":name,"Finding":"running","Details":"",
                                    "Status":"RUNNING","Delta":"NEW","Severity":"WARNING","Recommendation":f"New service {name} - verify"})
        if is_baseline:
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Services",
                            "Check":"Baseline","Finding":f"{len(current)} services catalogued",
                            "Details":"","Status":"OK","Delta":"BASELINE","Severity":"INFO","Recommendation":""})
        for svc in CRITICAL_LIN:
            if svc not in current:
                exists = run_cmd(f"systemctl list-unit-files {svc}.service 2>/dev/null | grep -c {svc}", timeout=5)
                try: svc_exists = int(exists.strip()) > 0
                except: svc_exists = False
                if svc_exists:
                    out2 = run_cmd(f"systemctl is-active {svc} 2>/dev/null")
                    if "active" not in out2:
                        critical_findings.append(f"SERVICE DOWN: {svc}")
                        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Services",
                                        "Check":svc,"Finding":"NOT running","Details":"Service is installed but stopped",
                                        "Status":"STOPPED","Delta":"","Severity":"WARNING",
                                        "Recommendation":f"Service {svc} is installed but not running - verify"})
    if is_baseline:
        g_baseline_services = current
    else:
        gone = g_baseline_services - current
        for s in gone:
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Services",
                            "Check":s,"Finding":"STOPPED","Details":"Was running previously",
                            "Status":"STOPPED","Delta":"GONE","Severity":"WARNING","Recommendation":f"Service {s} stopped - verify"})
        g_baseline_services = current
    status = "CRITICAL" if any("CRITICAL" in f for f in critical_findings) else \
             "WARNING" if critical_findings else "OK"
    findings_str = "; ".join(critical_findings) if critical_findings else \
                   (f"Baseline: {len(current)} services catalogued" if is_baseline else "No changes detected")
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"Services","Status":status,
                   "Critical Findings":findings_str,"Recommendation":"Investigate" if critical_findings else "Normal"})
    return rows_t, rows_e


def check_firewall(cycle):
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    if g_os_target == "windows":
        out = run_cmd("netsh advfirewall show allprofiles state")
        for line in out.splitlines():
            if "State" in line:
                state = line.split()[-1].strip()
                sev   = "OK" if state.upper() == "ON" else "CRITICAL"
                if sev == "CRITICAL": critical_findings.append(f"FIREWALL OFF: {line.strip()}")
                rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Firewall",
                                "Check":"Profile State","Finding":line.strip(),"Details":"",
                                "Status":state.upper(),"Delta":"","Severity":sev,
                                "Recommendation":"" if sev=="OK" else "Enable firewall immediately"})
    else:
        # Tolerant, multi-engine detection that works WITHOUT root by reading
        # config/service state, and uses live rules when root is available.
        try:    is_root = (os.geteuid() == 0)
        except Exception: is_root = False
        engines = []          # (name, active_bool_or_None, detail)

        # -- ufw: /etc/ufw/ufw.conf ENABLED= is authoritative and world-readable,
        #    even when `ufw status` refuses to run as non-root. --
        ufw_status = run_cmd("ufw status 2>/dev/null")            # root only
        ufw_conf   = run_cmd("grep -iE '^ENABLED' /etc/ufw/ufw.conf 2>/dev/null").lower()
        ufw_svc    = run_cmd("systemctl is-active ufw 2>/dev/null").strip()
        if "status: active" in ufw_status.lower():
            engines.append(("ufw", True, "ufw status: active"))
        elif "status: inactive" in ufw_status.lower():
            engines.append(("ufw", False, "ufw status: inactive"))
        elif ufw_conf:
            if "yes" in ufw_conf:
                engines.append(("ufw", True, "ufw.conf ENABLED=yes"))
            else:
                engines.append(("ufw", False, "ufw installed but disabled (ENABLED=no)"))
        elif ufw_svc == "active":
            engines.append(("ufw", None, "ufw.service active (run as root to confirm rules)"))

        # -- nftables --
        nft_rules = run_cmd("nft list ruleset 2>/dev/null")       # root only
        nft_svc   = run_cmd("systemctl is-active nftables 2>/dev/null").strip()
        if nft_rules.strip():
            n = len([l for l in nft_rules.splitlines() if l.strip().startswith(("tcp","udp","ip ","ct ","meta"))])
            engines.append(("nftables", True, f"nftables ruleset loaded (~{n} rule lines)"))
        elif nft_svc == "active":
            engines.append(("nftables", True, "nftables.service active"))

        # -- firewalld --
        if run_cmd("systemctl is-active firewalld 2>/dev/null").strip() == "active":
            engines.append(("firewalld", True, "firewalld active"))

        # -- raw iptables (root): more than the 3 default policy chains = real rules --
        ipt = run_cmd("iptables -S 2>/dev/null")                  # root only
        if ipt.strip():
            ipt_rules = [l for l in ipt.splitlines() if l.startswith("-A")]
            # docker/libvirt inject NAT rules; count host filter rules separately
            host_rules = [l for l in ipt_rules if not l.startswith(("-A DOCKER", "-A LIBVIRT"))]
            if host_rules:
                engines.append(("iptables", True, f"{len(host_rules)} iptables rule(s) active"))

        active_engines   = [e for e in engines if e[1] is True]
        disabled_engines = [e for e in engines if e[1] is False]
        unknown_engines  = [e for e in engines if e[1] is None]

        if active_engines:
            names  = ", ".join(e[0] for e in active_engines)
            detail = " | ".join(e[2] for e in active_engines)
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Firewall",
                            "Check":"Firewall Status","Finding":f"Active ({names})",
                            "Details":detail[:200],"Status":"ACTIVE","Delta":"","Severity":"OK",
                            "Recommendation":""})
        elif disabled_engines:
            detail = " | ".join(e[2] for e in disabled_engines)
            critical_findings.append("FIREWALL INSTALLED BUT DISABLED")
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Firewall",
                            "Check":"Firewall Status","Finding":"Installed but DISABLED",
                            "Details":detail[:200],"Status":"DISABLED","Delta":"","Severity":"WARNING",
                            "Recommendation":"A firewall is present but not enforcing - enable it (e.g. sudo ufw enable)"})
        elif unknown_engines:
            detail = " | ".join(e[2] for e in unknown_engines)
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Firewall",
                            "Check":"Firewall Status","Finding":"Present - state needs root to confirm",
                            "Details":detail[:200],"Status":"UNKNOWN","Delta":"","Severity":"INFO",
                            "Recommendation":"Run IT Angel as root to read live firewall rules"})
        elif not is_root:
            # No config/service signal AND we cannot read live rules: don't cry wolf.
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Firewall",
                            "Check":"Firewall Status","Finding":"Undetermined (no root)",
                            "Details":"No ufw/nftables/firewalld service or config detected without root",
                            "Status":"UNKNOWN","Delta":"","Severity":"INFO",
                            "Recommendation":"Run IT Angel as root to verify host firewall"})
        else:
            critical_findings.append("NO ACTIVE FIREWALL DETECTED")
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Firewall",
                            "Check":"Firewall Status","Finding":"No firewall detected",
                            "Details":"Checked ufw, nftables, firewalld and iptables - none active",
                            "Status":"NONE","Delta":"","Severity":"WARNING",
                            "Recommendation":"Enable ufw or configure iptables/nftables"})
    status = "CRITICAL" if any(r.get("Severity") == "CRITICAL" for r in rows_t) else \
             "WARNING" if any(r.get("Severity") == "WARNING" for r in rows_t) else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"Firewall","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else "Firewall active",
                   "Recommendation":"Enable firewall" if critical_findings else "Normal"})
    return rows_t, rows_e

def check_usb_devices(cycle):
    global g_baseline_usb
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    if not hasattr(check_usb_devices, "_known"):
        check_usb_devices._known        = {}
    if not hasattr(check_usb_devices, "_gone_persist"):
        check_usb_devices._gone_persist = {}
    USB_CLASS = {
        0x00:"Device", 0x01:"Audio", 0x02:"Communications (CDC)",
        0x03:"HID (Human Interface)", 0x05:"Physical", 0x06:"Image",
        0x07:"Printer", 0x08:"Mass Storage", 0x09:"USB Hub",
        0x0A:"CDC-Data", 0x0B:"Smart Card", 0x0D:"Content Security",
        0x0E:"Video", 0x0F:"Personal Healthcare", 0x10:"Audio/Video",
        0xDC:"Diagnostic", 0xE0:"Wireless Controller", 0xEF:"Miscellaneous",
        0xFE:"Application Specific", 0xFF:"Vendor Specific",
    }
    USB_SPEED = {
        0x0100:"USB 1.0 (1.5 Mbps)", 0x0110:"USB 1.1 (12 Mbps)",
        0x0200:"USB 2.0 (480 Mbps)", 0x0201:"USB 2.0 (480 Mbps)",
        0x0300:"USB 3.0 (5 Gbps)",   0x0310:"USB 3.1 (10 Gbps)",
        0x0320:"USB 3.2 (20 Gbps)",
    }
    def _dev_key(d):
        return f"{d['vid']}:{d['pid']}:{d['com_port']}:{d['mac']}:{d['sn']}"
    def _usb_extra(vid, pid):
        info = {}
        try:
            import usb.core as _usbc
            import usb.util as _usbu
        except Exception:
            return info
        try:
            dev = _usbc.find(idVendor=vid, idProduct=pid)
            if dev is None:
                return info
            try:    info["manufacturer"]  = _usbu.get_string(dev, dev.iManufacturer)  if dev.iManufacturer  else "N/A"
            except Exception: info["manufacturer"]  = "N/A"
            try:    info["product_name"]  = _usbu.get_string(dev, dev.iProduct)       if dev.iProduct       else "N/A"
            except Exception: info["product_name"]  = "N/A"
            try:    info["serial_number"] = _usbu.get_string(dev, dev.iSerialNumber)  if dev.iSerialNumber  else "N/A"
            except Exception: info["serial_number"] = "N/A"
            bcd = getattr(dev, "bcdUSB", None)
            info["usb_version"] = USB_SPEED.get(bcd, f"USB bcdUSB=0x{bcd:04X}" if bcd else "Unknown")
            cls = getattr(dev, "bDeviceClass", 0xFF)
            info["device_class"] = USB_CLASS.get(cls, f"Class 0x{cls:02X}")
        except Exception:
            pass
        return info
    # lsusb name map (Linux): resolves human-readable names from the usb.ids
    # database even for hubs/controllers whose USB string descriptors are empty.
    lsusb_names = {}
    if os.name != 'nt':
        for line in run_cmd("lsusb 2>/dev/null").splitlines():
            if "ID " not in line:
                continue
            rest = line.split("ID ", 1)[1].strip()      # e.g. "0e8d:0608 MediaTek Inc. Wireless_Device"
            vp   = rest[:9].lower()
            name = rest[9:].strip()
            if ":" in vp:
                lsusb_names[vp] = name or vp

    def _resolve_name(vid, pid, extra, cls):
        """Never return 'N/A'/'Unknown': fall back lsusb -> class -> vid:pid."""
        prod = (extra.get("product_name") or "").strip()
        if prod and prod.upper() != "N/A":
            return prod
        vp = f"{vid:04x}:{pid:04x}"
        if vp in lsusb_names:
            return lsusb_names[vp]
        return USB_CLASS.get(cls, f"USB Device {vid:04X}:{pid:04X}")

    current   = {}
    serial_vp = set()
    try:
        import serial.tools.list_ports as _lp
        for port in _lp.comports():
            # Skip legacy motherboard UARTs (/dev/ttyS0..31, ttyAMA*, etc.):
            # they are always-present kernel phantoms, NOT connected devices.
            # A genuine USB-serial adapter has a VID/PID or "USB" in its hwid.
            hwid_up = (port.hwid or "").upper()
            if not port.vid and not port.pid and "USB" not in hwid_up:
                continue
            vid = port.vid or 0
            pid = port.pid or 0
            extra = _usb_extra(vid, pid)
            nm = (port.description or "").strip()
            if not nm or nm.lower() == "n/a":
                nm = _resolve_name(vid, pid, extra, 0xFF)
            d = {
                "dev_type": "Serial/COM",
                "name":     nm,
                "com_port": port.device,
                "vid":      f"0x{vid:04X}" if vid else "N/A",
                "pid":      f"0x{pid:04X}" if pid else "N/A",
                "mfr":      extra.get("manufacturer") or port.manufacturer or "N/A",
                "sn":       extra.get("serial_number") or port.serial_number or "N/A",
                "mac":      "N/A",
                "hwid":     port.hwid or "N/A",
                "usb_ver":  extra.get("usb_version", "Unknown"),
            }
            current[_dev_key(d)] = d
            serial_vp.add((vid, pid))
    except Exception:
        pass
    try:
        USB_NET_KW = ("usb", "rndis", "gadget", "android", "tethering", "mobile")
        for iface, addrs in psutil.net_if_addrs().items():
            if not any(kw in iface.lower() for kw in USB_NET_KW):
                continue
            for addr in addrs:
                if addr.family == psutil.AF_LINK and addr.address not in ("", "00:00:00:00:00:00"):
                    d = {
                        "dev_type": "USB Network", "name": iface, "com_port": "N/A",
                        "vid": "N/A", "pid": "N/A", "mfr": "N/A", "sn": "N/A",
                        "mac": addr.address.upper(), "hwid": "N/A", "usb_ver": "Unknown",
                    }
                    current[_dev_key(d)] = d
    except Exception:
        pass
    try:
        import usb.core as _usbc
        all_devs = _usbc.find(find_all=True)
        if all_devs:
            for dev in all_devs:
                vid = dev.idVendor
                pid = dev.idProduct
                if (vid, pid) in serial_vp:
                    continue
                extra = _usb_extra(vid, pid)
                cls = getattr(dev, "bDeviceClass", 0xFF)
                bcd = getattr(dev, "bcdUSB", None)
                d = {
                    "dev_type": USB_CLASS.get(cls, "USB Device"),
                    "name":     _resolve_name(vid, pid, extra, cls),
                    "com_port": "N/A", "vid": f"0x{vid:04X}", "pid": f"0x{pid:04X}",
                    "mfr": extra.get("manufacturer", "N/A"), "sn": extra.get("serial_number", "N/A"),
                    "mac": "N/A", "hwid": f"{vid:04X}:{pid:04X}",
                    "usb_ver": USB_SPEED.get(bcd, f"bcdUSB 0x{bcd:04X}" if bcd else "Unknown"),
                }
                current[_dev_key(d)] = d
    except Exception:
        pass
    for k, d in current.items():
        check_usb_devices._known[k] = d
    current_keys  = set(current.keys())
    is_baseline   = len(g_baseline_usb) == 0
    new_keys      = current_keys - g_baseline_usb
    gone_keys     = g_baseline_usb - current_keys
    if not is_baseline:
        for k in gone_keys:
            check_usb_devices._gone_persist[k] = 10
    for k in new_keys:
        check_usb_devices._gone_persist.pop(k, None)
    for k in list(check_usb_devices._gone_persist.keys()):
        check_usb_devices._gone_persist[k] -= 1
        if check_usb_devices._gone_persist[k] <= 0:
            del check_usb_devices._gone_persist[k]
    g_baseline_usb = current_keys
    for k, d in current.items():
        is_new  = k in new_keys and not is_baseline
        is_hid  = "hid" in d["dev_type"].lower()
        is_ser  = d["com_port"] != "N/A"
        sev     = "WARNING" if is_new else "OK"
        dp = []
        if is_ser:            dp.append(f"Port: {d['com_port']}")
        if d["mac"] != "N/A": dp.append(f"MAC: {d['mac']}")
        if d["vid"] != "N/A": dp.append(f"VID:{d['vid']} PID:{d['pid']}")
        if d["mfr"] != "N/A": dp.append(f"Mfr: {d['mfr']}")
        if d["sn"]  != "N/A": dp.append(f"SN: {d['sn']}")
        if is_new:
            critical_findings.append(f"NEW HID: {d['name'][:45]}" if is_hid else f"NEW USB: {d['name'][:45]}")
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"USB Devices",
                        "Check":d["dev_type"],"Finding":d["name"][:70],"Details":" | ".join(dp),
                        "Status":"CONNECTED",
                        "Delta":"NEW" if is_new else ("BASELINE" if is_baseline else ""),
                        "Severity":sev,
                        "Recommendation":"Verify new USB - HID may be BadUSB" if is_new and is_hid else ("Verify new USB device" if is_new else "")})
    for k in check_usb_devices._gone_persist:
        d_gone    = check_usb_devices._known.get(k, {})
        gone_name = d_gone.get("name", k[:60])
        gone_type = d_gone.get("dev_type", "Device")
        gone_port = d_gone.get("com_port", "N/A")
        detail    = f"Port: {gone_port}" if gone_port != "N/A" else "Removed this session"
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"USB Devices",
                        "Check":gone_type,"Finding":gone_name[:70],"Details":detail,
                        "Status":"DISCONNECTED","Delta":"GONE","Severity":"WARNING","Recommendation":"USB device disconnected"})
    if not current and not check_usb_devices._gone_persist:
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"USB Devices",
                        "Check":"No devices","Finding":"None","Details":"",
                        "Status":"OK","Delta":"","Severity":"INFO","Recommendation":""})
    status = "WARNING" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"USB Devices","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else f"{len(current)} device(s) - no changes",
                   "Recommendation":"Verify new USB" if critical_findings else "Normal"})
    return rows_t, rows_e

def check_dns_gateway(cycle):
    ts = now_str(); rows_t = []; rows_e = []; findings = []
    if g_os_target == "windows":
        gw_out  = run_cmd("ipconfig | findstr /i \"Default Gateway\"")
        dns_out = run_cmd("ipconfig /all | findstr /i \"DNS Servers\"")
    else:
        gw_out  = run_cmd("ip route | grep default | awk '{print $3}'")
        dns_out = run_cmd("cat /etc/resolv.conf 2>/dev/null | grep -i nameserver | awk '{print $2}' | paste -sd' '")
        # systemd-resolved publishes a local stub (127.0.0.53/127.0.0.54) in
        # resolv.conf. That hides the real upstream servers, so resolve them.
        stub = (not dns_out) or all(x.startswith("127.0.0.5") for x in dns_out.split())
        if stub:
            real = run_cmd("resolvectl status 2>/dev/null | awk '/DNS Servers:/{$1=\"\";$2=\"\";print;exit}'").strip()
            if not real:
                real = run_cmd("resolvectl dns 2>/dev/null | awk '{for(i=2;i<=NF;i++) if($i ~ /[0-9]/) print $i}' | paste -sd' '").strip()
            if real:
                dns_out = f"{real}  (upstream; local stub {dns_out or '127.0.0.53'})"
    for check, out in [("Default Gateway", gw_out), ("DNS Servers", dns_out)]:
        val = out[:160] if out else "N/A"
        findings.append(f"{check}: {val[:60]}")
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"DNS & Gateway",
                        "Check":check,"Finding":val,"Details":"",
                        "Status":"INFO","Delta":"","Severity":"INFO","Recommendation":""})
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"DNS & Gateway","Status":"OK",
                   "Critical Findings":" | ".join(findings),"Recommendation":"Verify DNS matches expected servers"})
    return rows_t, rows_e


def check_scheduled_tasks(cycle):
    global g_baseline_tasks
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []; current = set()
    is_baseline = len(g_baseline_tasks) == 0
    if g_os_target == "windows":
        out = run_cmd('schtasks /query /fo LIST 2>nul | findstr /i "TaskName"', timeout=20)
        for line in out.splitlines():
            if "TaskName:" in line:
                tname = line.split("TaskName:")[-1].strip()
                if not tname: continue
                current.add(tname)
                is_new = tname not in g_baseline_tasks and not is_baseline
                if is_new:
                    critical_findings.append(f"NEW TASK: {tname[:50]}")
                    rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Scheduled Tasks",
                                    "Check":tname[:60],"Finding":"NEW - appeared this cycle","Details":"",
                                    "Status":"WARNING","Delta":"NEW","Severity":"WARNING","Recommendation":"Verify this new scheduled task"})
        if is_baseline:
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Scheduled Tasks",
                            "Check":"Baseline","Finding":f"{len(current)} tasks catalogued",
                            "Details":"Baseline established - only NEW tasks reported going forward",
                            "Status":"OK","Delta":"BASELINE","Severity":"INFO","Recommendation":""})
    else:
        out = run_cmd("crontab -l 2>/dev/null; ls /etc/cron.d/ 2>/dev/null; ls /etc/cron.daily/ 2>/dev/null")
        for line in out.splitlines():
            line = line.strip()
            if not line or line.startswith("#"): continue
            current.add(line)
            is_new = line not in g_baseline_tasks and not is_baseline
            if is_new:
                critical_findings.append(f"NEW CRON: {line[:50]}")
                rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Scheduled Tasks",
                                "Check":"cron","Finding":line[:80],"Details":"",
                                "Status":"WARNING","Delta":"NEW","Severity":"WARNING","Recommendation":"New cron entry - verify"})
        if is_baseline:
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Scheduled Tasks",
                            "Check":"Baseline","Finding":f"{len(current)} cron entries catalogued",
                            "Details":"Baseline established","Status":"OK","Delta":"BASELINE","Severity":"INFO","Recommendation":""})
    if is_baseline:
        g_baseline_tasks = current
    else:
        gone = g_baseline_tasks - current
        for t in gone:
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Scheduled Tasks",
                            "Check":t[:60],"Finding":"REMOVED since last cycle","Details":"",
                            "Status":"WARNING","Delta":"GONE","Severity":"WARNING","Recommendation":"Verify task removal is expected"})
        g_baseline_tasks = current
    status = "WARNING" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"Scheduled Tasks","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else
                   (f"Baseline: {len(current)} tasks catalogued" if is_baseline else "No new tasks detected"),
                   "Recommendation":"Verify new tasks" if critical_findings else "Normal"})
    return rows_t, rows_e


def check_local_network(cycle):
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    if g_os_target == "windows":
        out = run_cmd("arp -a")
    else:
        out = run_cmd("arp -n 2>/dev/null || ip neigh 2>/dev/null")
    current_devices = {}
    for line in out.splitlines():
        line = line.strip()
        if not line or "Interface" in line or "Address" in line: continue
        parts = line.split()
        if parts:
            ip = parts[0]
            if ip.count(".") == 3 and not ip.startswith("224.") and not ip.startswith("239."):
                current_devices[ip] = line[:80]
    if not hasattr(check_local_network, '_baseline'):
        check_local_network._baseline = set(current_devices.keys())
        new_ips = set()
    else:
        new_ips = set(current_devices.keys()) - check_local_network._baseline
        check_local_network._baseline = set(current_devices.keys())
    for ip in new_ips:
        if not ip.startswith("169.254."):
            critical_findings.append(f"NEW DEVICE: {ip}")
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"LAN Devices",
                            "Check":"New Device","Finding":ip,"Details":current_devices.get(ip,""),
                            "Status":"NEW","Delta":"NEW","Severity":"WARNING","Recommendation":f"New device {ip} joined the network - verify"})
    rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"LAN Devices",
                    "Check":"LAN Summary","Finding":f"{len(current_devices)} devices on network",
                    "Details":f"IPs: {', '.join(sorted(current_devices.keys())[:10])}",
                    "Status":"OK","Delta":"","Severity":"OK","Recommendation":""})
    status = "WARNING" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"LAN Devices","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else f"{len(current_devices)} devices - no new devices",
                   "Recommendation":"Verify new devices" if critical_findings else "Normal"})
    return rows_t, rows_e

def check_system_events(cycle):
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    if g_os_target == "windows":
        for evtid, desc, sev in [
            ("4625","Failed Login Attempts","CRITICAL"),
            ("4720","New User Account Created","CRITICAL"),
            ("7045","New Service Installed","WARNING"),
            ("4648","Logon with Explicit Credentials","WARNING"),
            ("4688","New Process Created","INFO"),
        ]:
            out = run_cmd(f'wevtutil qe Security /q:"*[System[EventID={evtid}]]" /c:5 /rd:true /f:text 2>nul | findstr /i "Date"', timeout=12)
            count = len([l for l in out.splitlines() if l.strip()])
            if count > 0 and sev in ("CRITICAL","WARNING"):
                critical_findings.append(f"Event {evtid} ({desc}): {count} recent")
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"System Events",
                            "Check":f"Event {evtid}","Finding":desc,"Details":f"Recent: {count}",
                            "Status":"FOUND" if count > 0 else "CLEAN","Delta":"",
                            "Severity":sev if count > 0 else "OK",
                            "Recommendation":f"Review event {evtid}" if count > 0 else ""})
    else:
        out = run_cmd("grep -i 'failed password\\|authentication failure' /var/log/auth.log 2>/dev/null | tail -10")
        count = len([l for l in out.splitlines() if l.strip()])
        if count > 0: critical_findings.append(f"Failed auth attempts: {count}")
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"System Events",
                        "Check":"Failed Logins","Finding":f"{count} recent","Details":out[:300] if out else "",
                        "Status":"FOUND" if count > 0 else "CLEAN","Delta":"",
                        "Severity":"CRITICAL" if count>5 else ("WARNING" if count>0 else "OK"),
                        "Recommendation":"Investigate brute force" if count>5 else ""})
        out2 = run_cmd("grep -i 'session opened for user root' /var/log/auth.log 2>/dev/null | tail -5")
        count2 = len([l for l in out2.splitlines() if l.strip()])
        if count2 > 0: critical_findings.append(f"Root sessions: {count2}")
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"System Events",
                        "Check":"Root Sessions","Finding":f"{count2} in log","Details":out2[:200] if out2 else "",
                        "Status":"FOUND" if count2>0 else "CLEAN","Delta":"",
                        "Severity":"WARNING" if count2>0 else "OK",
                        "Recommendation":"Verify root access is authorized" if count2>0 else ""})
        out3 = run_cmd("grep -i 'sudo:' /var/log/auth.log 2>/dev/null | tail -5")
        count3 = len([l for l in out3.splitlines() if l.strip()])
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"System Events",
                        "Check":"Sudo Usage","Finding":f"{count3} recent sudo events","Details":out3[:200] if out3 else "",
                        "Status":"INFO","Delta":"","Severity":"INFO","Recommendation":""})
    status = "CRITICAL" if any("CRITICAL" in str(r.get("Severity")) for r in rows_t) else \
             "WARNING" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"System Events","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else "No suspicious events",
                   "Recommendation":"Investigate" if critical_findings else "Normal"})
    return rows_t, rows_e


def check_open_shares(cycle):
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    if g_os_target == "windows":
        out = run_cmd("net share 2>nul")
        for line in out.splitlines()[2:]:
            if not line.strip() or "---" in line: continue
            if line.strip().lower().startswith("the "):
                continue
            parts = line.split()
            if not parts: continue
            name = parts[0]
            if len(name) > 40 or " " in name or name.lower() in ("the","this","there"):
                continue
            sev  = "WARNING" if name.upper() not in ("C$","IPC$","ADMIN$","PRINT$") else "INFO"
            if sev == "WARNING": critical_findings.append(f"OPEN SHARE: {name}")
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Network Shares",
                            "Check":name,"Finding":line.strip()[:80],"Details":"",
                            "Status":"SHARED","Delta":"","Severity":sev,
                            "Recommendation":"Verify share is intentional" if sev=="WARNING" else ""})
    else:
        found = 0
        # 1) NFS exports
        nfs = run_cmd("exportfs -v 2>/dev/null")
        for line in nfs.splitlines():
            line = line.strip()
            if not line: continue
            found += 1
            critical_findings.append(f"NFS EXPORT: {line[:40]}")
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Network Shares",
                            "Check":"NFS Export","Finding":line[:80],"Details":"exportfs -v",
                            "Status":"SHARED","Delta":"","Severity":"WARNING",
                            "Recommendation":"Verify NFS export is intentional"})
        # 2) Configured Samba shares (testparm reads smb.conf without needing a
        #    live connection - smbstatus alone shows only *connected* sessions).
        smb_running = (run_cmd("ss -tlnH 2>/dev/null | grep -c ':445'").strip() not in ("", "0") or
                       bool(run_cmd("pgrep -x smbd 2>/dev/null")))
        tp = run_cmd("testparm -s 2>/dev/null | grep -oP '^\\[\\K[^]]+'", timeout=8)
        for share in tp.splitlines():
            share = share.strip()
            if not share or share.lower() == "global": continue
            found += 1
            builtin = share.lower() in ("printers", "print$", "ipc$")
            sev = "INFO" if builtin else "WARNING"
            if not builtin: critical_findings.append(f"SMB SHARE: {share}")
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Network Shares",
                            "Check":f"SMB [{share}]","Finding":"Samba share exported",
                            "Details":"smbd listening on 445/139" if smb_running else "defined in smb.conf",
                            "Status":"SHARED","Delta":"","Severity":sev,
                            "Recommendation":"Verify share is intentional" if not builtin else ""})
        if found == 0:
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Network Shares",
                            "Check":"Shares","Finding":"No active shares detected","Details":"",
                            "Status":"OK","Delta":"","Severity":"OK","Recommendation":""})
    status = "WARNING" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"Network Shares","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else "No unexpected shares",
                   "Recommendation":"Verify shares" if critical_findings else "Normal"})
    return rows_t, rows_e


def check_startup_items(cycle):
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    if g_os_target == "windows":
        out = run_cmd('reg query "HKCU\\Software\\Microsoft\\Windows\\CurrentVersion\\Run" 2>nul')
        out += "\n"
        out += run_cmd('reg query "HKLM\\Software\\Microsoft\\Windows\\CurrentVersion\\Run" 2>nul')
        for line in out.splitlines():
            line = line.strip()
            if not line or line.startswith("HKEY"): continue
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Startup Items",
                            "Check":"Registry Run","Finding":line[:80],"Details":"",
                            "Status":"INFO","Delta":"","Severity":"INFO","Recommendation":""})
    else:
        out = run_cmd("systemctl list-unit-files --state=enabled --no-pager 2>/dev/null")
        current_units = set()
        for line in out.splitlines()[1:]:
            if not line.strip(): continue
            unit = line.split()[0] if line.split() else ""
            if unit: current_units.add(unit)
        if not hasattr(check_startup_items, '_baseline_lin'):
            check_startup_items._baseline_lin = set(current_units)
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Startup Items",
                            "Check":"Baseline","Finding":f"{len(current_units)} enabled units catalogued",
                            "Details":"","Status":"OK","Delta":"BASELINE","Severity":"INFO","Recommendation":""})
        else:
            new_units  = current_units - check_startup_items._baseline_lin
            gone_units = check_startup_items._baseline_lin - current_units
            for u in new_units:
                rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Startup Items",
                                "Check":"New Unit","Finding":u[:80],"Details":"",
                                "Status":"WARNING","Delta":"NEW","Severity":"WARNING","Recommendation":f"New enabled unit: {u}"})
            for u in gone_units:
                rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Startup Items",
                                "Check":"Removed Unit","Finding":u[:80],"Details":"",
                                "Status":"INFO","Delta":"GONE","Severity":"INFO","Recommendation":""})
            if not new_units and not gone_units:
                rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Startup Items",
                                "Check":"Status","Finding":"No changes","Details":"",
                                "Status":"OK","Delta":"","Severity":"OK","Recommendation":""})
            check_startup_items._baseline_lin = current_units
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"Startup Items","Status":"OK",
                   "Critical Findings":"Startup items catalogued","Recommendation":"Review for unexpected entries"})
    return rows_t, rows_e

def check_local_users(cycle):
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    if g_os_target == "windows":
        out = run_cmd("net user 2>nul")
        admins_out = run_cmd("net localgroup administrators 2>nul")
        admin_list = set(); in_members = False
        for line in admins_out.splitlines():
            if "---" in line: in_members = True; continue
            if in_members and line.strip() and "The command" not in line:
                admin_list.add(line.strip().lower())
        users_found = []
        for line in out.splitlines():
            line = line.strip()
            if not line or "User accounts" in line or "---" in line or "The command" in line: continue
            for u in line.split():
                if u: users_found.append(u)
        for u in users_found:
            is_admin = u.lower() in admin_list
            sev = "WARNING" if is_admin and u.lower() not in ("administrator","admin") else "OK"
            if is_admin: critical_findings.append(f"ADMIN USER: {u}")
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Local Users",
                            "Check":u,"Finding":"Admin" if is_admin else "Standard",
                            "Details":"Member of Administrators group" if is_admin else "",
                            "Status":"ADMIN" if is_admin else "STANDARD","Delta":"","Severity":sev,
                            "Recommendation":f"Verify {u} requires admin rights" if is_admin else ""})
    else:
        out = run_cmd("getent passwd | awk -F: '$7 !~ /nologin|false/ {print $1\":\"$3\":\"$6\":\"$7}'")
        sudo_out = run_cmd("getent group sudo wheel 2>/dev/null | cut -d: -f4")
        sudo_users = {u.strip() for line in sudo_out.splitlines() for u in line.split(",") if u.strip()}
        for line in out.splitlines():
            parts = line.split(":")
            if len(parts) < 4: continue
            uname, uid, home, shell = parts[0], parts[1], parts[2], parts[3]
            try: uid_int = int(uid)
            except: uid_int = -1
            is_root = uid == "0"; is_sudo = uname in sudo_users
            is_primary_user = uid_int >= 1000
            sev = "CRITICAL" if is_root and uname != "root" else ("WARNING" if is_sudo and not is_primary_user else "OK")
            if is_root and uname != "root": critical_findings.append(f"EXTRA ROOT: {uname}")
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Local Users",
                            "Check":uname,"Finding":f"UID:{uid}","Details":f"Shell:{shell} | Sudo:{is_sudo}",
                            "Status":"ROOT" if is_root else ("SUDO" if is_sudo else "NORMAL"),
                            "Delta":"","Severity":sev,
                            "Recommendation":f"Verify {uname} requires elevated access" if sev != "OK" else ""})
    status = "CRITICAL" if any("CRITICAL" in f for f in critical_findings) else "WARNING" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"Local Users","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else "No unexpected privileged users",
                   "Recommendation":"Audit admin accounts" if critical_findings else "Normal"})
    return rows_t, rows_e


def check_rdp_sessions(cycle):
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    if g_os_target == "windows":
        out = run_cmd("query session 2>nul")
        lines = out.splitlines()
        hdr = lines[0].upper() if lines else ""
        col_u  = hdr.find("USERNAME") if "USERNAME" in hdr else 19
        col_id = hdr.find(" ID")      if " ID"      in hdr else 38
        col_st = hdr.find("STATE")    if "STATE"    in hdr else 46
        col_u  = max(col_u, 0); col_id = max(col_id, 0); col_st = max(col_st, 0)
        for line in lines[1:]:
            if not line.strip(): continue
            session_name = line[:col_u].strip().lstrip(">").strip()
            username     = line[col_u:col_id].strip()
            state_raw    = line[col_st:col_st+12].strip() if col_st < len(line) else ""
            state        = state_raw.split()[0] if state_raw else "Unknown"
            if not username or username.isdigit():
                continue
            is_rdp    = "rdp" in session_name.lower() or "tcp" in session_name.lower()
            is_active = state.lower() == "active"
            sev = "WARNING" if (is_rdp and is_active) else "OK"
            if is_rdp and is_active:
                critical_findings.append(f"RDP SESSION: {username} on {session_name} [{state}]")
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"RDP Sessions",
                            "Check":session_name or "console","Finding":username,"Details":f"State:{state}",
                            "Status":state or "Unknown","Delta":"","Severity":sev,
                            "Recommendation":"Verify RDP session is authorized" if is_rdp and is_active else ""})
    else:
        found = 0
        seen  = set()
        # Primary source: loginctl (covers ssh, xrdp, console - even when utmp
        # is empty). "who | grep pts" alone misses xrdp/systemd remote logins.
        for s in get_loginctl_sessions():
            key = (s["user"], s["host"], s["session"])
            if key in seen: continue
            seen.add(key); found += 1
            is_remote = bool(s["remote"]) and s["host"] not in ('','::1','localhost','127.0.0.1')
            sev = "WARNING" if is_remote else "OK"
            label = "XRDP/RDP" if s["type"] in ("x11","wayland") and is_remote else ("SSH" if is_remote else "Local")
            if is_remote: critical_findings.append(f"{label} SESSION: {s['user']} from {s['host']}")
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"RDP Sessions",
                            "Check":f"{label} {s['user']}","Finding":s["host"] if is_remote else "local",
                            "Details":f"Session:{s['session']} Type:{s['type']} State:{s['state']}",
                            "Status":"REMOTE" if is_remote else "LOCAL","Delta":"","Severity":sev,
                            "Recommendation":"Verify remote session" if is_remote else ""})
        # Secondary source: who/pts (catches raw ptys not tracked by logind)
        out = run_cmd("who 2>/dev/null | grep pts")
        for line in out.splitlines():
            if not line.strip(): continue
            parts = line.split()
            uname = parts[0] if parts else "N/A"
            from_ = parts[-1].strip("()") if parts else "N/A"
            if (uname, from_) in {(k[0], k[1]) for k in seen}: continue
            is_remote = "." in from_ or ":" in from_
            sev = "WARNING" if is_remote else "OK"
            found += 1
            if is_remote: critical_findings.append(f"SSH SESSION: {uname} from {from_}")
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"RDP Sessions",
                            "Check":f"SSH {uname}","Finding":from_,"Details":"",
                            "Status":"REMOTE" if is_remote else "LOCAL","Delta":"","Severity":sev,
                            "Recommendation":"Verify remote SSH session" if is_remote else ""})
        if found == 0:
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"RDP Sessions",
                            "Check":"Remote Sessions","Finding":"None active","Details":"",
                            "Status":"OK","Delta":"","Severity":"OK","Recommendation":""})
    status = "WARNING" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"RDP Sessions","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else "No remote sessions",
                   "Recommendation":"Investigate" if critical_findings else "Normal"})
    return rows_t, rows_e

def check_privileged_processes(cycle):
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    EXPECTED_SYSTEM = {
        "System","System Idle Process","Memory Compression","Secure System","Registry",
        "smss.exe","csrss.exe","wininit.exe","winlogon.exe","services.exe","lsass.exe",
        "svchost.exe","spoolsv.exe","dwm.exe","sihost.exe","ctfmon.exe","taskhostw.exe",
        "conhost.exe","dllhost.exe","fontdrvhost.exe","RuntimeBroker.exe","audiodg.exe",
        "MsMpEng.exe","NisSrv.exe","SecurityHealthService.exe","WmiPrvSE.exe","msedge.exe",
    }
    EXPECTED_ROOT_LIN = {
        "systemd","kthreadd","ksoftirqd","kworker","rcu_sched","rcu_preempt",
        "migration","watchdog","cpuhp","idle_inject","kdevtmpfs","netns","kauditd",
        "khungtaskd","oom_reaper","writeback","kcompactd","ksmd","khugepaged",
        "cryptd","kintegrityd","kblockd","blkcg_punt_bio","tpm_dev_wq","ata_sff",
        "md","edac-poller","devfreq_wq","kswapd","ecryptfs-kthrea",
        "cfg80211","kstrp","zswap","charger_manager","usb-storage",
        "bioset","sshd","cron","crond","rsyslogd","agetty","NetworkManager",
        "wpa_supplicant","polkitd","accounts-daemon","systemd-journald",
        "systemd-udevd","systemd-logind","systemd-resolved","systemd-timesyncd",
        "systemd-networkd","systemd-journal","systemd-userdbd","systemd-userwor",
        "dbus-daemon","avahi-daemon","bluetoothd","udisksd",
        "packagekitd","thermald","irqbalance","rtkit-daemon","upowerd","colord",
        "gdm","gdm3","lightdm","Xorg","Xwayland","gvfsd","gvfs-udisks2",
        "snapd","dockerd","containerd","VBoxService","vmtoolsd","open-vm-tools",
        "mysqld","apache2","nginx","php-fpm","postgres","redis","mongod",
        "python3","python","bash","sh","dash","zsh","fish",
        "networkd-dispat","unattended-upgr","ModemManager","switcheroo-cont",
        "nvidia-smi","nvidia-persistenced","i2cdetect",
        "rustdesk","nordvpnd","openvpn","wireguard",
        "pool_workqueue_release","rcu_tasks_kthread","rcu_tasks_rude_kthread",
        "rcu_tasks_trace_kthread","rcu_exp_gp_kthread_worker","kdamond.0",
        "hwrng","haveged","smartd","watchdogd","psimon","fusermount3",
        "sudo","grep","ps","awk","cat","sed","cut","sort","tr",
        "head","tail","find","ip","ss","systemctl","journalctl",
        "iptables","nft","ufw","loginctl","hostnamectl","apt","dpkg",
        "tshark","dumpcap","krfcommd","obexd",
        "xfce4-session","xfce4-panel","xfdesktop","xfwm4","xfsettingsd",
        "dconf-service","dbus-launch","dbus-run-session","dbus-broker","dbus-broker-lau",
        "pipewire","pipewire-pulse","wireplumber","pulseaudio",
        "polkit","udisks2","geoclue","fwupd","cupsd","cups-browsed",
        "dhclient","dhcpcd","clamd","freshclam","gnome-keyring-d",
    }
    KERNEL_THREAD_PREFIXES = (
        "kworker/","kthread","ksoftirqd/","migration/","watchdog/",
        "cpuhp/","idle_inject/","kcompactd","kswapd",
        "irq/","card0-crtc","jbd2/","kdamond.","scsi_eh_",
        "rcu_exp_","rcu_tasks_","pool_workqueue",
    )
    if g_os_target == "windows":
        out = run_cmd('tasklist /v /fo csv 2>nul', timeout=15)
        system_procs = []; unexpected = []
        for line in out.splitlines()[1:]:
            line = line.strip().strip('"')
            if not line: continue
            parts = [p.strip('"') for p in line.split('","')]
            if len(parts) < 7: continue
            pname = parts[0]; user = parts[6]
            if "SYSTEM" in user.upper() or "NT AUTHORITY" in user.upper():
                system_procs.append(pname)
                if pname not in EXPECTED_SYSTEM:
                    unexpected.append(pname)
                    critical_findings.append(f"UNEXPECTED SYSTEM: {pname}")
                    rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Privileged Processes",
                                    "Check":pname,"Finding":f"{pname} (PID {parts[1]})",
                                    "Details":f"Account: {user[:40]} | Not in known SYSTEM list",
                                    "Status":"UNEXPECTED","Delta":"","Severity":"WARNING",
                                    "Recommendation":f"Investigate {pname} running as SYSTEM"})
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Privileged Processes",
                        "Check":"SYSTEM Summary","Finding":f"{len(system_procs)} total | {len(unexpected)} unexpected",
                        "Details":"Only unexpected processes listed above",
                        "Status":"WARNING" if unexpected else "OK","Delta":"",
                        "Severity":"WARNING" if unexpected else "OK",
                        "Recommendation":"Investigate" if unexpected else ""})
    else:
        out = run_cmd("ps -eo user,pid,comm --no-headers 2>/dev/null | grep '^root'")
        root_procs = []; unexpected = []
        for line in out.splitlines():
            parts = line.split()
            if len(parts) < 3: continue
            pname = parts[2]; pid = parts[1]
            root_procs.append(pname)
            is_kernel = any(pname.startswith(p) for p in KERNEL_THREAD_PREFIXES)
            if pname not in EXPECTED_ROOT_LIN and not is_kernel:
                unexpected.append(pname)
                critical_findings.append(f"UNEXPECTED ROOT: {pname}")
                rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Privileged Processes",
                                "Check":pname,"Finding":f"PID:{pid} User:root",
                                "Details":"Not in known root process list",
                                "Status":"UNEXPECTED","Delta":"","Severity":"WARNING",
                                "Recommendation":f"Investigate {pname} running as root"})
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Privileged Processes",
                        "Check":"Root Summary","Finding":f"{len(root_procs)} total | {len(unexpected)} unexpected",
                        "Details":"Only unexpected listed above",
                        "Status":"WARNING" if unexpected else "OK","Delta":"",
                        "Severity":"WARNING" if unexpected else "OK",
                        "Recommendation":"Investigate" if unexpected else ""})
    status = "WARNING" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"Privileged Processes","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else "No unexpected privileged processes",
                   "Recommendation":"Investigate" if critical_findings else "Normal"})
    return rows_t, rows_e

def check_recent_software(cycle):
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    if not hasattr(check_recent_software, '_sw_baseline'):
        check_recent_software._sw_baseline = None
    if g_os_target == "windows":
        found = {}
        for reg_path in [
            r"HKLM\Software\Microsoft\Windows\CurrentVersion\Uninstall",
            r"HKLM\Software\WOW6432Node\Microsoft\Windows\CurrentVersion\Uninstall",
        ]:
            out = run_cmd(f'reg query "{reg_path}" /s /v InstallDate 2>nul', timeout=20)
            current_key = ""
            for line in out.splitlines():
                line = line.strip()
                if line.startswith("HKEY"):
                    current_key = line; continue
                if "InstallDate" not in line: continue
                parts = line.split()
                date_str = parts[-1] if parts else ""
                if len(date_str) != 8 or not date_str.isdigit(): continue
                try:
                    days_ago = (datetime.datetime.now() - datetime.datetime.strptime(date_str, "%Y%m%d")).days
                except Exception:
                    continue
                if days_ago > 7 or not current_key: continue
                name_out = run_cmd(f'reg query "{current_key}" /v DisplayName 2>nul', timeout=5)
                name = ""
                for nl in name_out.splitlines():
                    if "DisplayName" in nl and "REG_SZ" in nl:
                        idx = nl.find("REG_SZ")
                        if idx >= 0: name = nl[idx + 6:].strip()
                if not name: continue
                if name not in found:
                    date_fmt = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:8]}"
                    found[name] = (date_fmt, days_ago)
        if check_recent_software._sw_baseline is None:
            check_recent_software._sw_baseline = set(found.keys())
        for name, (date_fmt, days_ago) in sorted(found.items(), key=lambda x: x[1][1]):
            is_new = name not in check_recent_software._sw_baseline
            sev    = "WARNING" if is_new else "INFO"
            label  = f"{name[:55]} ({date_fmt}, {days_ago}d ago)"
            if is_new: critical_findings.append(f"NEW SOFTWARE: {name[:50]}")
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Recent Software",
                            "Check":"Recent Install","Finding":label[:80],
                            "Details":f"Installed {days_ago} day(s) ago",
                            "Status":"NEW" if is_new else "RECENT","Delta":"NEW" if is_new else "",
                            "Severity":sev,"Recommendation":"Verify this installation" if is_new else ""})
        check_recent_software._sw_baseline = set(found.keys())
        if not rows_t:
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Recent Software",
                            "Check":"Recent Installs","Finding":"None in last 7 days","Details":"",
                            "Status":"OK","Delta":"","Severity":"OK","Recommendation":""})
    else:
        out = run_cmd("grep 'install ' /var/log/dpkg.log 2>/dev/null | tail -10")
        if not out: out = run_cmd("grep 'Installed:' /var/log/yum.log 2>/dev/null | tail -10")
        for line in out.splitlines():
            if not line.strip(): continue
            critical_findings.append(f"PACKAGE: {line[:50]}")
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Recent Software",
                            "Check":"Package","Finding":line[:80],"Details":"",
                            "Status":"RECENT","Delta":"","Severity":"WARNING","Recommendation":"Verify this package is authorized"})
        if not rows_t:
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Recent Software",
                            "Check":"Packages","Finding":"None detected","Details":"",
                            "Status":"OK","Delta":"","Severity":"OK","Recommendation":""})
    status = "WARNING" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"Recent Software","Status":status,
                   "Critical Findings":f"{len(critical_findings)} recent install(s)" if critical_findings else "No recent installs",
                   "Recommendation":"Audit recent software" if critical_findings else "Normal"})
    return rows_t, rows_e


def check_system_file_integrity(cycle):
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    import hashlib
    if g_os_target == "windows":
        critical_files = [
            r"C:\Windows\System32\drivers\etc\hosts",
            r"C:\Windows\System32\drivers\etc\services",
        ]
    else:
        critical_files = [
            "/etc/hosts","/etc/passwd","/etc/shadow",
            "/etc/sudoers","/etc/crontab","/etc/ssh/sshd_config",
        ]
    for fpath in critical_files:
        try:
            if not os.path.exists(fpath):
                rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"File Integrity",
                                "Check":os.path.basename(fpath),"Finding":"NOT FOUND","Details":"",
                                "Status":"MISSING","Delta":"","Severity":"WARNING",
                                "Recommendation":f"Expected file {fpath} is missing"})
                continue
            stat_r = os.stat(fpath)
            mtime  = datetime.datetime.fromtimestamp(stat_r.st_mtime).strftime("%Y-%m-%d %H:%M:%S")
            hours_ago = (time.time() - stat_r.st_mtime) / 3600
            with open(fpath,"rb") as f: chunk = f.read(4096)
            fhash = hashlib.md5(chunk).hexdigest()[:16]
            sev = "CRITICAL" if hours_ago < 1 else "WARNING" if hours_ago < 24 else "OK"
            if sev in ("CRITICAL","WARNING"):
                critical_findings.append(f"MODIFIED: {os.path.basename(fpath)} ({hours_ago:.1f}h ago)")
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"File Integrity",
                            "Check":os.path.basename(fpath),"Finding":mtime,
                            "Details":f"Size:{stat_r.st_size}B | MD5:{fhash} | Modified:{hours_ago:.1f}h ago",
                            "Status":"MODIFIED" if sev!="OK" else "OK","Delta":"","Severity":sev,
                            "Recommendation":f"Verify change to {fpath}" if sev!="OK" else ""})
        except PermissionError:
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"File Integrity",
                            "Check":os.path.basename(fpath),"Finding":"ACCESS DENIED","Details":"",
                            "Status":"DENIED","Delta":"","Severity":"INFO","Recommendation":""})
        except: pass
    status = "CRITICAL" if any("CRITICAL" in f for f in critical_findings) else \
             "WARNING" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"File Integrity","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else "No critical file modifications",
                   "Recommendation":"Investigate modified system files" if critical_findings else "Normal"})
    return rows_t, rows_e

def check_security_events_detail(cycle):
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    if g_os_target == "windows":
        event_checks = [
            ("4625","Failed Login","CRITICAL",3),("4740","Account Lockout","CRITICAL",1),
            ("4720","User Account Created","CRITICAL",1),("4726","User Account Deleted","CRITICAL",1),
            ("4732","Added to Admin Group","CRITICAL",1),("4733","Removed from Admin Group","WARNING",1),
            ("4648","Explicit Credential Logon","WARNING",1),("4719","Audit Policy Changed","CRITICAL",1),
            ("4698","Scheduled Task Created","WARNING",1),("4702","Scheduled Task Modified","WARNING",1),
            ("7045","New Service Installed","WARNING",1),("1102","Audit Log Cleared","CRITICAL",1),
        ]
        for evtid, desc, sev, threshold in event_checks:
            out = run_cmd(f'wevtutil qe Security /q:"*[System[EventID={evtid}]]" /c:{threshold+2} /rd:true /f:text 2>nul | findstr /i "Date"', timeout=10)
            count = len([l for l in out.splitlines() if l.strip()])
            if count >= threshold and sev != "INFO":
                critical_findings.append(f"Event {evtid} ({desc}): {count}")
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Security Events Detail",
                            "Check":f"Event {evtid}","Finding":desc,"Details":f"Recent occurrences: {count}",
                            "Status":"FOUND" if count>0 else "CLEAN","Delta":"",
                            "Severity":sev if count>=threshold else "OK",
                            "Recommendation":f"Investigate Event {evtid}" if count>=threshold else ""})
    else:
        for cmd_str, desc, sev in [
            ("grep -c 'Failed password' /var/log/auth.log 2>/dev/null", "SSH Failed Passwords","CRITICAL"),
            ("grep -c 'Invalid user' /var/log/auth.log 2>/dev/null",    "Invalid SSH Users",   "WARNING"),
            ("grep -c 'Accepted password' /var/log/auth.log 2>/dev/null","SSH Successful Logins","INFO"),
            ("grep -c 'sudo:' /var/log/auth.log 2>/dev/null",           "Sudo Usage Events",   "INFO"),
            ("grep -c 'useradd\\|userdel\\|usermod' /var/log/auth.log 2>/dev/null","User Account Changes","CRITICAL"),
        ]:
            out = run_cmd(cmd_str, timeout=8).strip()
            try: count = int(out)
            except: count = 0
            if count > 0 and sev in ("CRITICAL","WARNING"):
                critical_findings.append(f"{desc}: {count}")
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Security Events Detail",
                            "Check":desc,"Finding":str(count),"Details":"",
                            "Status":"FOUND" if count>0 else "CLEAN","Delta":"",
                            "Severity":sev if count>0 else "OK",
                            "Recommendation":f"Review {desc}" if (count>0 and sev!="INFO") else ""})
    status = "CRITICAL" if any("CRITICAL" in f for f in critical_findings) else \
             "WARNING" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"Security Events Detail","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else "No critical security events",
                   "Recommendation":"Investigate" if critical_findings else "Normal"})
    return rows_t, rows_e


def check_extended_ports(cycle):
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []; warnings = []
    # HIGH-CONFIDENCE: ports almost never used by legitimate software. Their
    # presence is a strong IOC -> CRITICAL.
    MALICIOUS = {
        1337:"Backdoor", 4444:"Metasploit/Meterpreter", 31337:"Backdoor (Elite)",
        31338:"Backdoor", 12345:"NetBus", 54321:"Backdoor", 65535:"Backdoor",
        6666:"IRC Bot", 6667:"IRC C2", 6668:"IRC C2", 6669:"IRC C2",
        1234:"Backdoor", 3460:"RAT", 4899:"Radmin RAT", 45560:"Miner", 45700:"Miner",
        9030:"Tor Relay", 9050:"Tor SOCKS", 9051:"Tor Control",
    }
    # DUAL-USE: frequently these are legitimate services (proxies, dashboards,
    # dev servers, VNC, mining pools, alt-HTTP). Flag with CONTEXT, never blind
    # CRITICAL - that would false-positive on almost every real server.
    DUAL_USE = {
        1080:"SOCKS proxy", 1099:"Java RMI", 2222:"Alt-SSH", 3333:"Mining/dev",
        3004:"Mining", 3008:"Mining", 5555:"ADB/dev", 5900:"VNC", 5901:"VNC", 5902:"VNC",
        7000:"Dev/Cassandra", 7777:"Game/dev", 8080:"HTTP-alt/proxy", 8181:"HTTP-alt",
        8443:"HTTPS-alt", 8888:"HTTP-alt/Jupyter", 9001:"Supervisor/Tor", 9090:"Cockpit/Prometheus",
        9999:"Dev/alt", 14444:"Mining", 14433:"Mining",
    }
    # Well-known legitimate owners: if a dual-use port belongs to one of these,
    # it is treated as benign (INFO). Keeps the tool quiet on normal servers.
    KNOWN_PROCS = {
        "node","node.js","python","python3","java","nginx","apache2","httpd","caddy",
        "docker-proxy","dockerd","containerd","containerd-shim","podman","conmon",
        "cockpit-ws","cockpit-tls","prometheus","grafana-server","node_exporter",
        "code","code-server","vscode-server","jupyter","jupyter-lab","gunicorn","uvicorn",
        "nginx:","php-fpm","ruby","rails","dotnet","gitlab-runner","jenkins","haproxy",
        "traefik","envoy","kong","supervisord","x11vnc","tigervnc","vncserver","Xvnc",
        "rustdesk","tailscaled","nordvpn","openvpn","wireguard","sshd","systemd-resolve",
    }
    try:
        is_root = (os.geteuid() == 0)
    except Exception:
        is_root = False
    try:
        flagged = {}
        for c in psutil.net_connections(kind='inet'):
            # Build (port, is_local_listen) candidates from both endpoints.
            cand = []
            if c.laddr: cand.append((c.laddr.port, getattr(c.laddr, "ip", ""), True))
            if c.raddr: cand.append((c.raddr.port, getattr(c.raddr, "ip", ""), False))
            for p, ipaddr, is_local in cand:
                if p in MALICIOUS:  cat, base = MALICIOUS[p], "MALICIOUS"
                elif p in DUAL_USE: cat, base = DUAL_USE[p], "DUAL_USE"
                else: continue
                try:    pname = psutil.Process(c.pid).name() if c.pid else "N/A"
                except Exception: pname = "N/A"
                key = (p, pname, base)
                if key in flagged: continue
                flagged[key] = True
                loopback = ipaddr.startswith("127.") or ipaddr in ("::1",)
                known    = pname.lower() in KNOWN_PROCS

                if base == "MALICIOUS":
                    sev = "CRITICAL"; label = f"{cat} port {p}"
                    critical_findings.append(f"{cat} port {p} by {pname}")
                    rec = f"INVESTIGATE: {cat} on port {p} (process {pname})"
                elif known:
                    # Dual-use port owned by a recognised service -> benign.
                    sev = "INFO"; label = f"{cat} port {p} (known: {pname})"
                    rec = ""
                elif loopback:
                    # Bound to localhost only: not externally exposed.
                    sev = "INFO"; label = f"{cat} port {p} (localhost only)"
                    rec = ""
                elif pname == "N/A" and not is_root:
                    # Cannot attribute the process without root: flag softly.
                    sev = "WARNING"; label = f"{cat} port {p} (owner unknown)"
                    warnings.append(f"{cat} port {p} - run as root to identify owner")
                    rec = f"Dual-use port {p} exposed - re-run as root to identify the process"
                else:
                    sev = "WARNING"; label = f"{cat} port {p} by {pname}"
                    warnings.append(f"{cat} port {p} by {pname}")
                    rec = f"Verify dual-use port {p} ({cat}) owned by {pname}"

                rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Extended Port Scan",
                                "Check":f"Port {p}","Finding":label,
                                "Details":f"Process:{pname} | Addr:{ipaddr or '?'} | Status:{c.status}",
                                "Status":c.status,"Delta":"","Severity":sev,"Recommendation":rec})
        if not flagged:
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Extended Port Scan",
                            "Check":"All ports","Finding":"No suspicious ports detected","Details":"",
                            "Status":"CLEAN","Delta":"","Severity":"OK","Recommendation":""})
    except Exception as ex:
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Extended Port Scan",
                        "Check":"Error","Finding":str(ex),"Details":"",
                        "Status":"ERROR","Delta":"","Severity":"WARNING","Recommendation":"Run as administrator/root"})
    status = "CRITICAL" if critical_findings else ("WARNING" if warnings else "OK")
    summary = "; ".join(critical_findings) if critical_findings else \
              ("; ".join(warnings) if warnings else "No suspicious ports")
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"Extended Port Scan","Status":status,
                   "Critical Findings":summary[:200],
                   "Recommendation":"Investigate immediately" if critical_findings else
                                    ("Verify flagged dual-use ports" if warnings else "Normal")})
    return rows_t, rows_e


def get_local_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        if not ip.startswith("169.254.") and not ip.startswith("127."):
            return ip
    except Exception:
        pass
    try:
        for iface, addrs in psutil.net_if_addrs().items():
            stats = psutil.net_if_stats().get(iface)
            if not stats or not stats.isup:
                continue
            for a in addrs:
                if a.family == socket.AF_INET:
                    ip = a.address
                    if (not ip.startswith("127.") and not ip.startswith("169.254.") and not ip.startswith("0.")):
                        return ip
    except Exception:
        pass
    return "N/A"

def get_live_connections():
    results = []
    try:
        conns = psutil.net_connections(kind='inet')
        for c in conns:
            if c.status not in ('ESTABLISHED', 'LISTEN', 'CLOSE_WAIT', 'TIME_WAIT'):
                continue
            try:    pname = psutil.Process(c.pid).name() if c.pid else "System"
            except: pname = "System"
            src = f"{c.laddr.ip}:{c.laddr.port}" if c.laddr else ""
            dst = f"{c.raddr.ip}:{c.raddr.port}" if c.raddr else "-"
            results.append({"src":src,"dst":dst,
                "proto":"TCP" if c.type == socket.SOCK_STREAM else "UDP",
                "status":c.status,"proc":pname})
    except Exception:
        pass
    order = {"ESTABLISHED": 0, "CLOSE_WAIT": 1, "TIME_WAIT": 2, "LISTEN": 3}
    results.sort(key=lambda x: (order.get(x["status"], 9), x["dst"]))
    return results[:20]


def get_lan_devices():
    devices = []; seen = set()
    local_ip = get_local_ip()
    if os.name == 'nt':
        subnet = ".".join(local_ip.split(".")[:3]) if local_ip and local_ip != "N/A" else ""
        alive_ips = set()
        def ping_one(ip):
            try:
                r = subprocess.run(["ping","-n","1","-w","300",ip], capture_output=True, timeout=1)
                if r.returncode == 0: alive_ips.add(ip)
            except: pass
        if subnet:
            threads = []
            for i in range(1, 255):
                ip = f"{subnet}.{i}"
                if ip == local_ip: alive_ips.add(ip); continue
                t = threading.Thread(target=ping_one, args=(ip,), daemon=True)
                threads.append(t); t.start()
            for t in threads: t.join(timeout=0.4)
        try:
            out = run_cmd("arp -a", timeout=3)
            for line in out.splitlines():
                line = line.strip()
                if not line or "Interface" in line or "Address" in line: continue
                parts = line.split()
                if len(parts) < 2: continue
                ip = parts[0]; mac = parts[1] if len(parts) > 1 else "?"
                if ip.count(".") != 3: continue
                if ip.startswith("224.") or ip.startswith("239."): continue
                if ip.endswith(".255") or ip.startswith("255."): continue
                if ip in seen: continue
                seen.add(ip)
                devices.append({"ip": ip, "mac": mac, "alive": ip in alive_ips or ip == local_ip})
        except: pass
    else:
        try:
            out = run_cmd("ip neigh show 2>/dev/null", timeout=3)
            for line in out.splitlines():
                line = line.strip()
                if not line: continue
                parts = line.split()
                if not parts: continue
                ip = parts[0]
                if ip.count(".") != 3: continue
                if ip.startswith("224.") or ip.startswith("239."): continue
                if ip.endswith(".255") or ip.startswith("255."): continue
                if ip in seen: continue
                mac = "?"
                if "lladdr" in parts:
                    idx = parts.index("lladdr")
                    if idx + 1 < len(parts): mac = parts[idx + 1]
                state = parts[-1].upper() if parts else ""
                if state in ("FAILED", "INCOMPLETE"): continue
                seen.add(ip)
                devices.append({"ip": ip, "mac": mac, "alive": True})
        except: pass
    if local_ip and local_ip not in seen:
        devices.insert(0, {"ip": local_ip, "mac": "local", "alive": True})
    return devices


def get_net_stats():
    try:
        s = psutil.net_io_counters()
        return s.bytes_sent, s.bytes_recv, s.packets_sent, s.packets_recv
    except Exception:
        return 0, 0, 0, 0


def get_local_subnet():
    ip = get_local_ip()
    if ip and ip != "N/A":
        parts = ip.split(".")
        if len(parts) == 4:
            return ".".join(parts[:3])
    return "192.168.1"


def get_connection_bytes():
    activity = {}
    try:
        conns = psutil.net_connections(kind='inet')
        for c in conns:
            if c.status != 'ESTABLISHED' or not c.raddr:
                continue
            dst_ip = c.raddr.ip
            if dst_ip not in activity:
                activity[dst_ip] = 0
            activity[dst_ip] += 1
    except Exception:
        pass
    return activity


def capture_tshark_background(cycle_num, duration_sec, result_holder, done_event):
    if not g_tshark_path:
        done_event.set()
        return
    MAX_PKTS = 500
    iface = get_tshark_interface()
    cmd = [
        g_tshark_path, "-i", iface, "-a", f"duration:{duration_sec}", "-c", str(MAX_PKTS),
        "-T", "fields", "-e", "frame.time_epoch", "-e", "ip.src", "-e", "ip.dst",
        "-e", "_ws.col.Protocol", "-e", "_ws.col.Info",
        "-E", "separator=\t", "-E", "quote=n", "-E", "occurrence=f", "-l",
    ]
    rows = []
    try:
        proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.DEVNULL, text=True, bufsize=1)
        try:
            for line in proc.stdout:
                if g_stop_event.is_set():
                    break
                line = line.rstrip()
                if not line:
                    continue
                parts = line.split("\t")
                try:
                    epoch = float(parts[0].strip())
                    ts_str = datetime.datetime.fromtimestamp(epoch).strftime("%Y-%m-%d %H:%M:%S")
                except Exception:
                    ts_str = now_str()
                src   = parts[1].strip() if len(parts) > 1 else ""
                dst   = parts[2].strip() if len(parts) > 2 else ""
                proto = parts[3].strip() if len(parts) > 3 else ""
                info  = parts[4].strip()[:100] if len(parts) > 4 else ""
                if src or dst:
                    row = {"Timestamp": ts_str, "Cycle": cycle_num, "Source": src,
                           "Destination": dst, "Protocol": proto, "Info": info}
                    rows.append(row)
                    result_holder.append(row)
        except Exception:
            pass
        finally:
            try:
                proc.kill()
                proc.wait(timeout=5)
            except Exception:
                pass
    except Exception:
        pass
    if rows:
        save_traffic_to_excel(rows)
    done_event.set()


def save_traffic_to_excel(traffic_rows):
    if not traffic_rows:
        return
    with g_excel_lock:
        try:
            wb     = openpyxl.load_workbook(g_excel_path)
            ws_net = wb["Network Traffic"]
            next_n = ws_net.max_row + 1
            for i, row in enumerate(traffic_rows):
                r   = next_n + i
                alt = (r % 2 == 0)
                vals = [row.get("Timestamp",""), row.get("Cycle",""), row.get("Source",""),
                        row.get("Destination",""), row.get("Protocol",""), row.get("Info","")]
                for col, val in enumerate(vals, 1):
                    cell           = ws_net.cell(row=r, column=col, value=val)
                    cell.border    = thin_border
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    cell.fill      = FILL_ALT if alt else FILL_WHITE
                    cell.font      = FONT_BLACK
                ws_net.row_dimensions[r].height = 16
            wb.save(g_excel_path)
        except Exception:
            pass

def check_defender_status(cycle):
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    if g_os_target == "windows":
        out = run_cmd('powershell -NoProfile -Command "Get-MpComputerStatus | Select-Object AntivirusEnabled,RealTimeProtectionEnabled,TamperProtectionSource,AntivirusSignatureAge | ConvertTo-Csv -NoTypeInformation" 2>nul', timeout=15)
        csv_lines = [l.strip() for l in out.splitlines() if l.strip()]
        if len(csv_lines) >= 2:
            hdrs = [h.strip('"') for h in csv_lines[0].split(',')]
            vals = [v.strip('"') for v in csv_lines[1].split(',')]
            data = dict(zip(hdrs, vals))
            av_on  = data.get("AntivirusEnabled","").upper() == "TRUE"
            rtp_on = data.get("RealTimeProtectionEnabled","").upper() == "TRUE"
            tamper = data.get("TamperProtectionSource","")
            try: sig_age = int(data.get("AntivirusSignatureAge","0"))
            except: sig_age = 0
            for check, val, is_crit, is_warn, rec in [
                ("Antivirus Enabled",   str(av_on),  not av_on, False, "Enable Windows Defender AV"),
                ("Real-Time Protection",str(rtp_on), not rtp_on, False, "Enable real-time protection"),
                ("Signature Age (days)",str(sig_age), False, sig_age > 3, "Update Defender signatures"),
                ("Tamper Protection",   tamper, False, tamper == "", "Enable Tamper Protection"),
            ]:
                sev = "CRITICAL" if is_crit else ("WARNING" if is_warn else "OK")
                if is_crit: critical_findings.append(f"DEFENDER: {check} = {val}")
                rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Defender / AV",
                                "Check":check,"Finding":val,"Details":"","Status":sev,"Delta":"","Severity":sev,
                                "Recommendation":rec if (is_crit or is_warn) else ""})
        else:
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Defender / AV",
                            "Check":"Defender","Finding":"Unable to query","Details":out[:80],
                            "Status":"WARNING","Delta":"","Severity":"WARNING",
                            "Recommendation":"Run as administrator or check Defender service"})
    else:
        for tool, cmd_str in [("ClamAV","clamscan --version 2>/dev/null"),("ufw","ufw status 2>/dev/null | head -1")]:
            out = run_cmd(cmd_str, timeout=5)
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Defender / AV",
                            "Check":tool,"Finding":out[:60] if out else "Not found","Details":"",
                            "Status":"INFO","Delta":"","Severity":"INFO","Recommendation":""})
    status = "CRITICAL" if any(r.get("Severity") == "CRITICAL" for r in rows_t) else \
             "WARNING" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"Defender / AV","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else "Defender active",
                   "Recommendation":"Fix Defender issues" if critical_findings else "Normal"})
    return rows_t, rows_e


def check_windows_update(cycle):
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    if g_os_target == "windows":
        out = run_cmd('powershell -NoProfile -Command "Get-HotFix | Sort-Object InstalledOn -Descending | Select-Object -First 1 | Select-Object HotFixID,InstalledOn | ConvertTo-Csv -NoTypeInformation" 2>nul', timeout=15)
        csv_lines = [l.strip() for l in out.splitlines() if l.strip()]
        if len(csv_lines) >= 2:
            hdrs = [h.strip('"') for h in csv_lines[0].split(',')]
            vals = [v.strip('"') for v in csv_lines[1].split(',')]
            data = dict(zip(hdrs, vals))
            hfid = data.get("HotFixID","N/A"); inst = data.get("InstalledOn","N/A")
            try:
                inst_dt  = datetime.datetime.strptime(inst.split(" ")[0], "%m/%d/%Y")
                days_ago = (datetime.datetime.now() - inst_dt).days
            except Exception:
                days_ago = -1
            is_stale = days_ago > 30
            sev = "WARNING" if is_stale else "OK"
            if is_stale: critical_findings.append(f"STALE: Last patch {days_ago}d ago ({hfid})")
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Windows Update",
                            "Check":"Last Patch","Finding":f"{hfid} ({inst})",
                            "Details":f"{days_ago}d since last update" if days_ago >= 0 else "",
                            "Status":sev,"Delta":"","Severity":sev,
                            "Recommendation":"Run Windows Update" if is_stale else ""})
        else:
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Windows Update",
                            "Check":"Last Patch","Finding":"Unable to query","Details":"",
                            "Status":"INFO","Delta":"","Severity":"INFO","Recommendation":""})
    else:
        out = run_cmd("apt list --upgradable 2>/dev/null | wc -l", timeout=10)
        try: count = max(0, int(out.strip()) - 1)
        except: count = 0
        sev = "WARNING" if count > 10 else "OK"
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Windows Update",
                        "Check":"Pending Updates","Finding":f"{count} packages upgradable","Details":"",
                        "Status":sev,"Delta":"","Severity":sev,
                        "Recommendation":"Run apt upgrade" if sev == "WARNING" else ""})
    status = "WARNING" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"Windows Update","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else "System up to date",
                   "Recommendation":"Update system" if critical_findings else "Normal"})
    return rows_t, rows_e


def check_bitlocker_secureboot(cycle):
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    if g_os_target == "windows":
        bl_out = run_cmd('powershell -NoProfile -Command "Get-BitLockerVolume | Select-Object MountPoint,VolumeStatus,ProtectionStatus | ConvertTo-Csv -NoTypeInformation" 2>nul', timeout=15)
        for line in bl_out.splitlines()[1:]:
            parts = [p.strip('"') for p in line.split(',')]
            if len(parts) < 3: continue
            mount, vol_status, prot = parts[0], parts[1], parts[2]
            protected = "FullyEncrypted" in vol_status or prot == "On"
            sev = "OK" if protected else "WARNING"
            if not protected and mount.upper() == "C:":
                critical_findings.append("C: drive NOT encrypted")
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"BitLocker / SecureBoot",
                            "Check":f"BitLocker {mount}","Finding":vol_status,"Details":f"Protection: {prot}",
                            "Status":prot,"Delta":"","Severity":sev,
                            "Recommendation":"Enable BitLocker on system drive" if not protected and mount.upper() == "C:" else ""})
        if not bl_out.strip():
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"BitLocker / SecureBoot",
                            "Check":"BitLocker","Finding":"Not available or access denied","Details":"",
                            "Status":"INFO","Delta":"","Severity":"INFO","Recommendation":""})
        sb_out = run_cmd('powershell -NoProfile -Command "Confirm-SecureBootUEFI" 2>nul', timeout=10)
        sb_val = sb_out.strip(); sb_on = sb_val.upper() == "TRUE"
        sb_sev = "OK" if sb_on else ("WARNING" if sb_val else "INFO")
        if not sb_on and sb_val: critical_findings.append("Secure Boot disabled")
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"BitLocker / SecureBoot",
                        "Check":"Secure Boot","Finding":sb_val if sb_val else "N/A","Details":"",
                        "Status":"ON" if sb_on else ("OFF" if sb_val else "N/A"),"Delta":"","Severity":sb_sev,
                        "Recommendation":"Enable Secure Boot in BIOS" if not sb_on and sb_val else ""})
    else:
        sb = run_cmd("mokutil --sb-state 2>/dev/null", timeout=5)
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"BitLocker / SecureBoot",
                        "Check":"Secure Boot","Finding":sb[:60] if sb else "N/A","Details":"",
                        "Status":"INFO","Delta":"","Severity":"INFO","Recommendation":""})
        enc = run_cmd("lsblk -o FSTYPE | grep -c crypt", timeout=5)
        try: enc_cnt = int(enc.strip())
        except: enc_cnt = 0
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"BitLocker / SecureBoot",
                        "Check":"Disk Encryption","Finding":"Encrypted" if enc_cnt else "None detected","Details":"",
                        "Status":"OK" if enc_cnt else "INFO","Delta":"","Severity":"OK" if enc_cnt else "INFO","Recommendation":""})
    status = "WARNING" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"BitLocker / SecureBoot","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else "Encryption/SecureBoot OK",
                   "Recommendation":"Enable encryption/SecureBoot" if critical_findings else "Normal"})
    return rows_t, rows_e

def check_audit_policy(cycle):
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    if g_os_target == "windows":
        CRITICAL_SUBS = {
            "credential validation": ("Account Logon > Credential Validation","Password/hash checks"),
            "logon": ("Logon/Logoff > Logon","User sign-in"),
            "logoff": ("Logon/Logoff > Logoff","User sign-outs"),
            "account lockout": ("Logon/Logoff > Account Lockout","Brute force indicator"),
            "special logon": ("Logon/Logoff > Special Logon","Admin logons"),
            "process creation": ("Detailed Tracking > Process Creation","Malware detection"),
            "audit policy change": ("Policy Change > Audit Policy Change","Audit settings modified"),
            "user account management": ("Account Management > User Account Management","User changes"),
            "security group management": ("Account Management > Security Group Management","Group changes"),
            "sensitive privilege use": ("Privilege Use > Sensitive Privilege Use","Mimikatz indicator"),
            "security state change": ("System > Security State Change","Security subsystem"),
        }
        KNOWN_SETTINGS = ("success and failure", "success", "failure", "no auditing")
        out = run_cmd('auditpol /get /category:* 2>nul', timeout=15)
        matched = {}
        for line in out.splitlines():
            stripped = line.strip()
            if not stripped: continue
            low = stripped.lower()
            if not line.startswith(" ") and not any(s in low for s in KNOWN_SETTINGS):
                continue
            for s in KNOWN_SETTINGS:
                if low.endswith(s):
                    sub_name = stripped[:-len(s)].strip().lower()
                    if sub_name in CRITICAL_SUBS:
                        matched[sub_name] = stripped[-len(s):]
                    break
        for sub_key, (label, description) in CRITICAL_SUBS.items():
            setting = matched.get(sub_key, "Not Found")
            is_off  = setting.lower() in ("no auditing", "not found")
            sev     = "CRITICAL" if is_off else "OK"
            if is_off: critical_findings.append(f"AUDIT OFF: {label.split(' > ')[-1]}")
            short_name = label.split(" > ")[-1]
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Audit Policy",
                            "Check":label,"Finding":f"{short_name}: {setting}","Details":description,
                            "Status":sev,"Delta":"","Severity":sev,
                            "Recommendation":f"Enable '{short_name}' auditing (Success and Failure)" if is_off else ""})
        if not matched:
            rows_t = [{"Timestamp":ts,"Cycle":cycle,"Category":"Audit Policy",
                        "Check":"Audit Policy","Finding":"Unable to query - run as administrator","Details":"",
                        "Status":"WARNING","Delta":"","Severity":"WARNING","Recommendation":""}]
    else:
        out = run_cmd("auditctl -l 2>/dev/null", timeout=5)
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Audit Policy",
                        "Check":"Linux > auditctl rules","Finding":out[:80] if out else "No rules active",
                        "Details":"Kernel audit rules loaded by auditd","Status":"OK" if out else "WARNING",
                        "Delta":"","Severity":"OK" if out else "WARNING",
                        "Recommendation":"Configure auditd rules (auditctl -l)" if not out else ""})
    status = "CRITICAL" if any("CRITICAL" in f for f in critical_findings) else \
             "WARNING" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"Audit Policy","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else "All critical audit subcategories enabled",
                   "Recommendation":"Enable critical audit categories" if critical_findings else "Normal"})
    return rows_t, rows_e


def check_event_log_cleared(cycle):
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    if g_os_target == "windows":
        for evtid, log, desc in [("1102","Security","Security Log Cleared"),("104","System","System Log Cleared")]:
            out = run_cmd(f'wevtutil qe {log} /q:"*[System[EventID={evtid}]]" /c:5 /rd:true /f:text 2>nul', timeout=10)
            count = len([l for l in out.splitlines() if l.strip() and "Date" in l])
            sev = "CRITICAL" if count > 0 else "OK"
            if count > 0: critical_findings.append(f"LOG CLEARED: {desc} ({count} event(s))")
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Event Log Cleared",
                            "Check":desc,"Finding":f"{count} occurrence(s)","Details":out[:200] if count > 0 else "",
                            "Status":"CLEARED" if count > 0 else "CLEAN","Delta":"","Severity":sev,
                            "Recommendation":"Investigate who cleared the log" if count > 0 else ""})
    else:
        out = run_cmd("journalctl --list-boots 2>/dev/null | wc -l", timeout=5)
        try: boots = int(out.strip())
        except: boots = 0
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Event Log Cleared",
                        "Check":"Journal Boots","Finding":f"{boots} boot record(s)","Details":"",
                        "Status":"INFO","Delta":"","Severity":"INFO","Recommendation":""})
    status = "CRITICAL" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"Event Log Cleared","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else "No log clearing detected",
                   "Recommendation":"Investigate" if critical_findings else "Normal"})
    return rows_t, rows_e


def check_shadow_copies(cycle):
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    if g_os_target == "windows":
        vss_out  = run_cmd('vssadmin list shadows 2>nul', timeout=15)
        vss_cnt  = sum(1 for l in vss_out.splitlines() if "Shadow Copy ID" in l)
        del_out  = run_cmd('wevtutil qe System /q:"*[System[EventID=524]]" /c:5 /rd:true /f:text 2>nul | findstr "Date"', timeout=10)
        del_cnt  = len([l for l in del_out.splitlines() if l.strip()])
        sev_del  = "CRITICAL" if del_cnt > 0 else "OK"
        if del_cnt > 0: critical_findings.append(f"SHADOW COPIES DELETED: {del_cnt} event(s) - ransomware IOC")
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Shadow Copies",
                        "Check":"VSS Copies Present","Finding":f"{vss_cnt} shadow copy(ies)","Details":"",
                        "Status":"OK" if vss_cnt > 0 else "WARNING","Delta":"",
                        "Severity":"OK" if vss_cnt > 0 else "WARNING",
                        "Recommendation":"Create VSS snapshots for recovery" if vss_cnt == 0 else ""})
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Shadow Copies",
                        "Check":"VSS Deletion Events","Finding":f"{del_cnt} deletion event(s)",
                        "Details":del_out[:200] if del_cnt > 0 else "",
                        "Status":"CRITICAL" if del_cnt > 0 else "CLEAN","Delta":"","Severity":sev_del,
                        "Recommendation":"INVESTIGATE - possible ransomware" if del_cnt > 0 else ""})
    else:
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Shadow Copies",
                        "Check":"VSS","Finding":"Linux - not applicable","Details":"",
                        "Status":"INFO","Delta":"","Severity":"INFO","Recommendation":""})
    status = "CRITICAL" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"Shadow Copies","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else "Shadow copies intact",
                   "Recommendation":"Investigate ransomware IOC" if critical_findings else "Normal"})
    return rows_t, rows_e


def check_system_time(cycle):
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    if g_os_target == "windows":
        out = run_cmd('wevtutil qe Security /q:"*[System[EventID=4616]]" /c:5 /rd:true /f:text 2>nul | findstr "Date"', timeout=10)
        count = len([l for l in out.splitlines() if l.strip()])
        if count > 0: critical_findings.append(f"SYSTEM TIME CHANGED: {count} event(s)")
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"System Time",
                        "Check":"Time Change Events (4616)","Finding":f"{count} event(s)","Details":"",
                        "Status":"FOUND" if count > 0 else "CLEAN","Delta":"",
                        "Severity":"WARNING" if count > 0 else "OK",
                        "Recommendation":"Verify time change is expected" if count > 0 else ""})
        w32 = run_cmd('w32tm /query /status 2>nul', timeout=10)
        for line in w32.splitlines():
            if any(k in line for k in ("Stratum","Source","Last Sync")):
                rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"System Time",
                                "Check":"NTP Status","Finding":line.strip(),"Details":"",
                                "Status":"INFO","Delta":"","Severity":"INFO","Recommendation":""})
    else:
        out = run_cmd("timedatectl status 2>/dev/null | head -5", timeout=5)
        for line in out.splitlines():
            if line.strip():
                rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"System Time",
                                "Check":"Time Status","Finding":line.strip(),"Details":"",
                                "Status":"INFO","Delta":"","Severity":"INFO","Recommendation":""})
    status = "WARNING" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"System Time","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else "No time manipulation",
                   "Recommendation":"Investigate time change" if critical_findings else "Normal"})
    return rows_t, rows_e

def check_registry_autoruns(cycle):
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    if not hasattr(check_registry_autoruns, '_baseline'):
        check_registry_autoruns._baseline = {}
    if g_os_target == "windows":
        RUN_KEYS = [
            r"HKCU\Software\Microsoft\Windows\CurrentVersion\Run",
            r"HKCU\Software\Microsoft\Windows\CurrentVersion\RunOnce",
            r"HKLM\Software\Microsoft\Windows\CurrentVersion\Run",
            r"HKLM\Software\Microsoft\Windows\CurrentVersion\RunOnce",
        ]
        WINLOGON_KEY = r"HKLM\Software\Microsoft\Windows NT\CurrentVersion\Winlogon"
        IFEO_KEY = r"HKLM\Software\Microsoft\Windows NT\CurrentVersion\Image File Execution Options"
        current = {}
        for key in RUN_KEYS:
            out = run_cmd(f'reg query "{key}" 2>nul', timeout=8)
            for line in out.splitlines():
                line = line.strip()
                if not line or line.startswith("HKEY"): continue
                parts = line.split(None, 2)
                if len(parts) < 3: continue
                name, _, value = parts
                ekey = f"{key}\\{name}"
                current[ekey] = value[:120]
                is_new = ekey not in check_registry_autoruns._baseline and check_registry_autoruns._baseline
                if is_new: critical_findings.append(f"NEW AUTORUN: {name} = {value[:50]}")
                rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Registry Autoruns",
                                "Check":name[:40],"Finding":value[:70],"Details":f"Key: {key}",
                                "Status":"NEW" if is_new else "OK","Delta":"NEW" if is_new else "",
                                "Severity":"WARNING" if is_new else "INFO",
                                "Recommendation":"Verify this autorun entry" if is_new else ""})
        for val_name, expected_sub in [("Shell","explorer.exe"),("Userinit","userinit.exe")]:
            wl = run_cmd(f'reg query "{WINLOGON_KEY}" /v {val_name} 2>nul', timeout=8)
            for line in wl.splitlines():
                if val_name in line and "REG_SZ" in line:
                    parts = line.strip().split(None, 2)
                    if len(parts) < 3: continue
                    value = parts[2].strip()
                    is_sus = expected_sub.lower() not in value.lower()
                    sev = "CRITICAL" if is_sus else "OK"
                    if is_sus: critical_findings.append(f"WINLOGON HIJACK: {val_name} = {value[:60]}")
                    rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Registry Autoruns",
                                    "Check":f"Winlogon {val_name}","Finding":value[:70],"Details":"Winlogon key",
                                    "Status":"SUSPICIOUS" if is_sus else "OK","Delta":"","Severity":sev,
                                    "Recommendation":"INVESTIGATE Winlogon hijack" if is_sus else ""})
        ifeo = run_cmd(f'reg query "{IFEO_KEY}" /s /v Debugger 2>nul', timeout=12)
        for line in ifeo.splitlines():
            if "Debugger" in line and "REG_SZ" in line:
                parts = line.strip().split(None, 2)
                if len(parts) < 3: continue
                value = parts[2].strip()
                if not any(lg in value.lower() for lg in ("vsjitdebugger","drwatson","werfault")):
                    critical_findings.append(f"IFEO HIJACK: Debugger = {value[:60]}")
                    rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Registry Autoruns",
                                    "Check":"IFEO Debugger","Finding":value[:70],
                                    "Details":"Image File Execution Options hijack",
                                    "Status":"CRITICAL","Delta":"","Severity":"CRITICAL",
                                    "Recommendation":"INVESTIGATE: possible image hijack persistence"})
        check_registry_autoruns._baseline = current
        if not rows_t:
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Registry Autoruns",
                            "Check":"Autoruns","Finding":"No Run key entries found","Details":"",
                            "Status":"OK","Delta":"","Severity":"OK","Recommendation":""})
    else:
        for fpath in ["/etc/rc.local","/etc/profile"]:
            if os.path.isfile(fpath):
                rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Registry Autoruns",
                                "Check":"Startup Script","Finding":fpath,"Details":"",
                                "Status":"INFO","Delta":"","Severity":"INFO","Recommendation":""})
    status = "CRITICAL" if any("CRITICAL" in f for f in critical_findings) else \
             "WARNING" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"Registry Autoruns","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else "No suspicious autoruns",
                   "Recommendation":"Investigate" if critical_findings else "Normal"})
    return rows_t, rows_e


def check_wmi_subscriptions(cycle):
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    if g_os_target == "windows":
        for cls, fields in [("__EventFilter","Name,Query"),("__EventConsumer","Name"),("__FilterToConsumerBinding","Filter,Consumer")]:
            out = run_cmd(f'powershell -NoProfile -Command "Get-CimInstance -Namespace root/subscription -ClassName {cls} | Select-Object {fields} | ConvertTo-Csv -NoTypeInformation" 2>nul', timeout=12)
            lines = [l.strip() for l in out.splitlines() if l.strip()]
            for line in lines[1:]:
                vals = [v.strip('"') for v in line.split(',')]
                name = vals[0] if vals else ""
                if not name: continue
                critical_findings.append(f"WMI {cls}: {name[:50]}")
                rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"WMI Subscriptions",
                                "Check":cls,"Finding":name[:70],"Details":",".join(vals[1:])[:80],
                                "Status":"FOUND","Delta":"","Severity":"CRITICAL",
                                "Recommendation":f"Investigate WMI subscription: {name[:40]}"})
        if not rows_t:
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"WMI Subscriptions",
                            "Check":"WMI Persistence","Finding":"No subscriptions found","Details":"",
                            "Status":"CLEAN","Delta":"","Severity":"OK","Recommendation":""})
    else:
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"WMI Subscriptions",
                        "Check":"WMI","Finding":"Linux - not applicable","Details":"",
                        "Status":"INFO","Delta":"","Severity":"INFO","Recommendation":""})
    status = "CRITICAL" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"WMI Subscriptions","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else "No WMI subscriptions",
                   "Recommendation":"Investigate WMI persistence" if critical_findings else "Normal"})
    return rows_t, rows_e


def check_browser_extensions(cycle):
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    if not hasattr(check_browser_extensions, '_baseline'):
        check_browser_extensions._baseline = {}
    import json as _json
    appdata  = os.environ.get("LOCALAPPDATA","")
    home     = os.path.expanduser("~")
    if os.name == 'nt':
        browsers = {
            "Chrome": os.path.join(appdata, "Google","Chrome","User Data"),
            "Edge":   os.path.join(appdata, "Microsoft","Edge","User Data"),
            "Brave":  os.path.join(appdata, "BraveSoftware","Brave-Browser","User Data"),
        }
    else:
        browsers = {
            "Chrome":   os.path.join(home,".config","google-chrome"),
            "Chromium": os.path.join(home,".config","chromium"),
            "Brave":    os.path.join(home,".config","BraveSoftware","Brave-Browser"),
        }
    current_exts = {}
    for browser, data_dir in browsers.items():
        if not os.path.isdir(data_dir): continue
        for profile in ["Default"] + [f"Profile {i}" for i in range(1,6)]:
            ext_dir = os.path.join(data_dir, profile, "Extensions")
            if not os.path.isdir(ext_dir): continue
            try:
                for ext_id in os.listdir(ext_dir):
                    ext_path = os.path.join(ext_dir, ext_id)
                    if not os.path.isdir(ext_path): continue
                    name = ext_id
                    try:
                        for ver in os.listdir(ext_path):
                            mf = os.path.join(ext_path, ver, "manifest.json")
                            if os.path.isfile(mf):
                                with open(mf,"r",encoding="utf-8",errors="ignore") as f:
                                    name = _json.load(f).get("name",ext_id)[:60]
                                break
                    except Exception:
                        pass
                    key = f"{browser}/{profile}/{ext_id}"
                    current_exts[key] = name
                    is_new = key not in check_browser_extensions._baseline and check_browser_extensions._baseline
                    if is_new: critical_findings.append(f"NEW EXT: {browser} - {name[:40]}")
                    rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Browser Extensions",
                                    "Check":f"{browser}/{profile}","Finding":name[:70],"Details":f"ID: {ext_id}",
                                    "Status":"NEW" if is_new else "OK","Delta":"NEW" if is_new else "",
                                    "Severity":"WARNING" if is_new else "INFO",
                                    "Recommendation":"Verify new browser extension" if is_new else ""})
            except Exception:
                pass
    check_browser_extensions._baseline = current_exts
    if not rows_t:
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Browser Extensions",
                        "Check":"Extensions","Finding":"No supported browsers found","Details":"",
                        "Status":"INFO","Delta":"","Severity":"INFO","Recommendation":""})
    status = "WARNING" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"Browser Extensions","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else f"{len(current_exts)} extension(s) tracked",
                   "Recommendation":"Verify new extensions" if critical_findings else "Normal"})
    return rows_t, rows_e

def check_beaconing(cycle):
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    if not hasattr(check_beaconing, '_history'):
        check_beaconing._history = {}
    local_subnet = get_local_subnet()
    try:
        for c in psutil.net_connections(kind='inet'):
            if c.status != 'ESTABLISHED' or not c.raddr: continue
            dst = c.raddr.ip
            if dst.startswith(local_subnet+".") or dst.startswith("127."): continue
            check_beaconing._history.setdefault(dst, []).append(time.time())
    except Exception:
        pass
    now_t = time.time(); beacons = []
    for ip, ts_list in list(check_beaconing._history.items()):
        check_beaconing._history[ip] = [t for t in ts_list if now_t - t < 7200]
        if not check_beaconing._history[ip]:
            del check_beaconing._history[ip]; continue
        count = len(check_beaconing._history[ip])
        if count >= 3:
            beacons.append((ip, count))
            critical_findings.append(f"BEACON: {ip} seen {count} times")
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Beaconing",
                            "Check":"Periodic Connection","Finding":ip,"Details":f"Seen {count} times in 2h window",
                            "Status":"BEACON","Delta":"","Severity":"WARNING",
                            "Recommendation":f"Investigate periodic C2 connection to {ip}"})
    rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Beaconing",
                    "Check":"Beacon Monitor","Finding":f"{len(check_beaconing._history)} external IPs tracked",
                    "Details":f"{len(beacons)} potential beacon(s)",
                    "Status":"WARNING" if beacons else "OK","Delta":"",
                    "Severity":"WARNING" if beacons else "OK","Recommendation":""})
    status = "WARNING" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"Beaconing","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else "No beaconing detected",
                   "Recommendation":"Investigate beaconing IPs" if critical_findings else "Normal"})
    return rows_t, rows_e


def check_arp_spoofing(cycle):
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    if not hasattr(check_arp_spoofing, '_gw_mac'):
        check_arp_spoofing._gw_mac = None
    gw_ip = ""
    if g_os_target == "windows":
        gw_out = run_cmd("ipconfig | findstr /i \"Default Gateway\"", timeout=8)
        for line in gw_out.splitlines():
            if ":" in line:
                candidate = line.split(":")[-1].strip()
                if candidate.count(".") == 3 and not candidate.startswith("0."):
                    gw_ip = candidate; break
    else:
        raw = run_cmd("ip route show default 2>/dev/null", timeout=5)
        parts = raw.split()
        if "via" in parts:
            idx = parts.index("via")
            if idx + 1 < len(parts): gw_ip = parts[idx+1]
    if not gw_ip:
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"ARP / Gateway",
                        "Check":"Gateway","Finding":"Could not determine gateway","Details":"",
                        "Status":"INFO","Delta":"","Severity":"INFO","Recommendation":""})
        rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"ARP / Gateway","Status":"OK",
                       "Critical Findings":"No gateway","Recommendation":"Normal"})
        return rows_t, rows_e
    if os.name == 'nt':
        arp_out = run_cmd(f"arp -a {gw_ip}", timeout=5)
    else:
        arp_out = run_cmd(f"ip neigh show {gw_ip}", timeout=5)
    current_mac = ""
    for line in arp_out.splitlines():
        parts = line.split()
        if os.name == 'nt':
            if len(parts) >= 2 and "-" in parts[1] and len(parts[1]) == 17:
                current_mac = parts[1].replace("-",":").lower(); break
        else:
            if "lladdr" in parts:
                idx = parts.index("lladdr")
                if idx + 1 < len(parts):
                    current_mac = parts[idx+1].lower(); break
    if not current_mac:
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"ARP / Gateway",
                        "Check":f"Gateway {gw_ip}","Finding":"MAC not in ARP cache","Details":"",
                        "Status":"INFO","Delta":"","Severity":"INFO","Recommendation":""})
    elif check_arp_spoofing._gw_mac is None:
        check_arp_spoofing._gw_mac = current_mac
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"ARP / Gateway",
                        "Check":f"Gateway {gw_ip}","Finding":current_mac,"Details":"Baseline MAC captured",
                        "Status":"OK","Delta":"BASELINE","Severity":"INFO","Recommendation":""})
    elif current_mac != check_arp_spoofing._gw_mac:
        critical_findings.append(f"GATEWAY MAC CHANGED: {check_arp_spoofing._gw_mac} -> {current_mac}")
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"ARP / Gateway",
                        "Check":f"Gateway {gw_ip}","Finding":current_mac,
                        "Details":f"Was: {check_arp_spoofing._gw_mac} | Now: {current_mac}",
                        "Status":"CHANGED","Delta":"CHANGED","Severity":"CRITICAL",
                        "Recommendation":"POSSIBLE ARP SPOOFING / MITM ATTACK"})
        check_arp_spoofing._gw_mac = current_mac
    else:
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"ARP / Gateway",
                        "Check":f"Gateway {gw_ip}","Finding":current_mac,"Details":"MAC unchanged",
                        "Status":"OK","Delta":"","Severity":"OK","Recommendation":""})
    status = "CRITICAL" if any(r.get("Severity") == "CRITICAL" for r in rows_t) else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"ARP / Gateway","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else f"Gateway {gw_ip} MAC stable",
                   "Recommendation":"Investigate MITM" if critical_findings else "Normal"})
    return rows_t, rows_e

def check_process_lineage(cycle):
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    SUSPICIOUS_WIN = {
        ("winword.exe","powershell.exe"),("winword.exe","cmd.exe"),
        ("excel.exe","powershell.exe"),("excel.exe","cmd.exe"),
        ("outlook.exe","powershell.exe"),("outlook.exe","cmd.exe"),
        ("powershell.exe","mshta.exe"),("powershell.exe","wscript.exe"),
        ("cmd.exe","mshta.exe"),("mshta.exe","powershell.exe"),("regsvr32.exe","powershell.exe"),
    }
    SUSPICIOUS_LIN = {
        ("apache2","bash"),("nginx","bash"),("php-fpm","bash"),
        ("httpd","bash"),("python3","bash"),("perl","bash"),
    }
    if g_os_target == "windows":
        out = run_cmd('powershell -NoProfile -Command "Get-CimInstance Win32_Process | Select-Object Name,ProcessId,ParentProcessId | ConvertTo-Csv -NoTypeInformation" 2>nul', timeout=20)
        proc_map = {}; proc_parent = {}
        for line in out.splitlines()[1:]:
            parts = [p.strip('"') for p in line.split(',')]
            if len(parts) < 3: continue
            try:
                pid  = int(parts[1]); ppid = int(parts[2]) if parts[2] else 0
                proc_map[pid]    = parts[0].lower()
                proc_parent[pid] = ppid
            except Exception: continue
        for pid, name in proc_map.items():
            ppid  = proc_parent.get(pid, 0)
            pname = proc_map.get(ppid, "")
            if (pname, name) in SUSPICIOUS_WIN:
                critical_findings.append(f"CHAIN: {pname} -> {name} (PID {pid})")
                rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Process Lineage",
                                "Check":f"{pname} -> {name}","Finding":f"PID {pid}","Details":"Suspicious parent-child chain",
                                "Status":"SUSPICIOUS","Delta":"","Severity":"CRITICAL",
                                "Recommendation":f"Investigate: {pname} spawned {name}"})
    else:
        out = run_cmd("ps -eo pid,ppid,comm --no-headers 2>/dev/null", timeout=10)
        proc_map = {}; proc_parent = {}
        for line in out.splitlines():
            parts = line.split()
            if len(parts) < 3: continue
            try:
                pid = int(parts[0]); ppid = int(parts[1])
                proc_map[pid] = parts[2].lower(); proc_parent[pid] = ppid
            except Exception: continue
        for pid, name in proc_map.items():
            ppid  = proc_parent.get(pid, 0)
            pname = proc_map.get(ppid, "")
            if (pname, name) in SUSPICIOUS_LIN:
                critical_findings.append(f"WEB SHELL: {pname} -> {name}")
                rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Process Lineage",
                                "Check":f"{pname} -> {name}","Finding":f"PID {pid}","Details":"Web server spawned shell",
                                "Status":"SUSPICIOUS","Delta":"","Severity":"CRITICAL",
                                "Recommendation":"Investigate: web process spawned shell"})
    if not critical_findings:
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Process Lineage",
                        "Check":"Process Lineage","Finding":"No suspicious chains detected","Details":"",
                        "Status":"CLEAN","Delta":"","Severity":"OK","Recommendation":""})
    status = "CRITICAL" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"Process Lineage","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else "No suspicious chains",
                   "Recommendation":"Investigate chains" if critical_findings else "Normal"})
    return rows_t, rows_e


def check_ransomware_indicators(cycle):
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    RANSOM_EXTS = {
        ".locked",".encrypted",".enc",".crypt",".crypted",".crypto",
        ".cerber",".zepto",".locky",".wnry",".wncry",".wannacry",
        ".ryuk",".revil",".maze",".conti",".darkside",".blackcat",
        ".lockbit",".blackbasta",".hive",".ransomware",".pay2key",
    }
    RANSOM_NOTES = {"readme.txt","!decrypt","recover","ransom","_readme.txt",
                    "!!!readme!!!.txt","decrypt_instructions.txt","how to decrypt"}
    scan_dirs = []
    for d in ["Desktop","Documents","Downloads"]:
        p = os.path.join(os.path.expanduser("~"), d)
        if os.path.isdir(p): scan_dirs.append(p)
    suspicious = []; notes = []
    for scan_dir in scan_dirs:
        try:
            for fname in os.listdir(scan_dir):
                ext   = os.path.splitext(fname)[1].lower()
                flow  = fname.lower()
                fpath = os.path.join(scan_dir, fname)
                if ext in RANSOM_EXTS: suspicious.append(fpath)
                if any(n in flow for n in RANSOM_NOTES): notes.append(fpath)
        except Exception:
            pass
    if suspicious:
        critical_findings.append(f"RANSOMWARE FILES: {len(suspicious)} suspicious file(s)")
        for fp in suspicious[:5]:
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Ransomware IOC",
                            "Check":"Encrypted File","Finding":os.path.basename(fp)[:70],"Details":fp[:100],
                            "Status":"CRITICAL","Delta":"","Severity":"CRITICAL","Recommendation":"STOP all I/O - possible ransomware active"})
    if notes:
        critical_findings.append(f"RANSOM NOTES: {len(notes)} file(s)")
        for fp in notes[:3]:
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Ransomware IOC",
                            "Check":"Ransom Note","Finding":os.path.basename(fp)[:70],"Details":fp[:100],
                            "Status":"CRITICAL","Delta":"","Severity":"CRITICAL","Recommendation":"RANSOM NOTE FOUND - isolate system immediately"})
    if not critical_findings:
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Ransomware IOC",
                        "Check":"Ransomware Scan","Finding":"No indicators found",
                        "Details":f"Scanned: {', '.join(os.path.basename(d) for d in scan_dirs)}",
                        "Status":"CLEAN","Delta":"","Severity":"OK","Recommendation":""})
    status = "CRITICAL" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"Ransomware IOC","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else "No ransomware indicators",
                   "Recommendation":"ISOLATE SYSTEM" if critical_findings else "Normal"})
    return rows_t, rows_e


def check_webcam_mic(cycle):
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    if g_os_target == "windows":
        for dev_class, label in [("Camera","Webcam"),("AudioEndpoint","Microphone")]:
            out = run_cmd(f'powershell -NoProfile -Command "Get-PnpDevice -Class {dev_class} -PresentOnly | Select-Object Status,FriendlyName | ConvertTo-Csv -NoTypeInformation" 2>nul', timeout=12)
            for line in out.splitlines()[1:]:
                parts = [p.strip('"') for p in line.split(',')]
                if len(parts) < 2: continue
                dev_status, dev_name = parts[0], parts[1]
                rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Webcam / Mic",
                                "Check":label,"Finding":dev_name[:60],"Details":f"Status: {dev_status}",
                                "Status":dev_status,"Delta":"","Severity":"INFO","Recommendation":""})
        if not rows_t:
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Webcam / Mic",
                            "Check":"Camera/Mic","Finding":"No devices or access data found","Details":"",
                            "Status":"INFO","Delta":"","Severity":"INFO","Recommendation":""})
    else:
        out = run_cmd("ls /dev/video* 2>/dev/null", timeout=5)
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Webcam / Mic",
                        "Check":"Video Devices","Finding":out.strip() if out else "None","Details":"",
                        "Status":"INFO","Delta":"","Severity":"INFO","Recommendation":""})
    status = "WARNING" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"Webcam / Mic","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else "Camera/mic status logged",
                   "Recommendation":"Review" if critical_findings else "Normal"})
    return rows_t, rows_e

# --- LIVE WATCHMAN DASHBOARD (Textual, scrollable) -----------------------
_CHECK_LABELS = [
    "Active Connections", "Listening Ports", "Extended Port Scan", "Net Processes",
    "Privileged Processes", "Resources (CPU/RAM)", "Logged Users", "Local Users & Admins",
    "RDP / Remote Sessions", "Services", "Firewall", "USB Devices", "DNS & Gateway",
    "Scheduled Tasks", "LAN Devices (ARP)", "System Events", "Security Events Detail",
    "Network Shares", "Startup Items", "Recent Software", "File Integrity",
    "Defender / AV", "Windows Update", "BitLocker / SecureBoot",
    "Audit Policy", "Event Log Cleared", "Shadow Copies", "System Time",
    "Registry Autoruns", "WMI Subscriptions", "Browser Extensions",
    "Beaconing", "ARP / Gateway", "Process Lineage", "Ransomware IOC", "Webcam / Mic",
]
_DOT = {"OK": "[green]●[/]", "WARNING": "[yellow]●[/]",
        "CRITICAL": "[bold red]●[/]", "ERR": "[red]✗[/]", "-": "[dim]●[/]"}


class _WatchmanApp(App):
    CSS = """
    Screen { background: #0a0a0a; }
    #hdr { dock: top; height: 3; background: #07140a; border: heavy green; color: white; content-align: left middle; padding: 0 1; }
    #hdr.crit { border: heavy red; }
    #ftr { dock: bottom; height: 3; background: #07140a; border: heavy green; color: white; padding: 0 1; content-align: left middle; }
    .panel { border: round #2a6f8f; margin: 1 1 0 1; padding: 0 1; height: auto; }
    .alert { border: round red; margin: 1 1 0 1; padding: 0 1; height: auto; }
    """
    BINDINGS = [Binding("q", "stop", "Stop"), Binding("ctrl+c", "stop", "Stop")]

    def __init__(self, wait_seconds, start_time, duration_end, local_ip, local_subnet,
                 iface, cycle_num, cycle_tech, check_status, get_lan_now):
        super().__init__()
        self.wait_seconds = wait_seconds
        self.start_time   = start_time
        self.duration_end = duration_end
        self.local_ip     = local_ip
        self.local_subnet = local_subnet
        self.iface        = iface
        self.cycle_num    = cycle_num
        self.cycle_tech   = cycle_tech
        self.check_status = check_status
        self.get_lan_now  = get_lan_now
        self.prev_sent    = 0
        self.prev_recv    = 0
        self.activity_log = {}
        self.baseline_lan = set()

    def compose(self):
        yield Static(id="hdr")
        with VerticalScroll():
            yield Static(id="checks",   classes="panel")
            yield Static(id="alerts",   classes="alert")
            yield Static(id="conn",     classes="panel")
            yield Static(id="usb",      classes="panel")
            yield Static(id="sessions", classes="panel")
            yield Static(id="procs",    classes="panel")
            yield Static(id="ports",    classes="panel")
            yield Static(id="lan",      classes="panel")
        yield Static(id="ftr")

    def on_mount(self):
        try:
            s, r, _, _ = get_net_stats()
            self.prev_sent, self.prev_recv = s, r
        except Exception:
            pass
        try:
            self.baseline_lan = {d["ip"] for d in self.get_lan_now()}
        except Exception:
            self.baseline_lan = set()
        self._render_security()
        self.set_interval(2.0, self.tick)
        self.tick()

    def _rows_for(self, *cats):
        return [r for r in self.cycle_tech if r.get("Category") in cats]

    def _text_panel(self, title, *cats):
        rows = self._rows_for(*cats)
        if not rows:
            return "[bold cyan]%s[/]\n\n  [dim](no findings this cycle)[/]" % title
        lines = []
        for r in rows[:60]:
            sev = r.get("Severity", "OK")
            col = {"CRITICAL": "bold red", "WARNING": "yellow"}.get(sev, "white")
            fnd = str(r.get("Finding", ""))[:64]
            det = str(r.get("Details", ""))
            extra = ("  - " + det[:24]) if det else ""
            lines.append("  [%s]%s[/]%s" % (col, fnd, extra))
        return "[bold cyan]%s[/]\n\n%s" % (title, "\n".join(lines))

    def _render_security(self):
        cells = []
        for lbl in _CHECK_LABELS:
            st = self.check_status.get(lbl, "-")
            cells.append("%s %s" % (_DOT.get(st, _DOT["-"]), ("%-22s" % lbl)))
        grid = "\n".join("   ".join(cells[i:i+3]) for i in range(0, len(cells), 3))
        self.query_one("#checks", Static).update(
            "[bold cyan]SECURITY CHECKS[/]  [dim](cycle %s)[/]\n\n%s" % (self.cycle_num, grid))
        alerts = [r for r in self.cycle_tech if r.get("Severity") in ("CRITICAL", "WARNING")]
        if alerts:
            lines = []
            for r in alerts:
                sev = r.get("Severity")
                tag = "[bold red]CRITICAL[/]" if sev == "CRITICAL" else "[yellow]WARNING [/]"
                fnd = str(r.get("Finding", ""))[:70]
                lines.append("  %s [bold]%s:[/] %s" % (tag, r.get("Category", ""), fnd))
            body = "\n".join(lines)
        else:
            body = "  [green]No active alerts - all clear.[/]"
        self.query_one("#alerts", Static).update(
            "[bold red]ACTIVE ALERTS (%d)[/]\n\n%s" % (len(alerts), body))
        self.query_one("#sessions", Static).update(
            self._text_panel("SESSIONS / USERS", "Logged Users", "RDP Sessions"))
        self.query_one("#procs", Static).update(
            self._text_panel("PROCESSES (privileged / network)", "Privileged Processes", "Net Processes"))
        self.query_one("#ports", Static).update(
            self._text_panel("LISTENING / EXPOSED PORTS", "Listening Ports", "Extended Port Scan"))

    def tick(self):
        elapsed = time.time() - self.start_time
        left    = max(0, self.wait_seconds - int(elapsed))
        if g_stop_event.is_set() or left <= 0:
            self.exit(); return
        m, s = divmod(left, 60)
        try:
            cur_sent, cur_recv, _, _ = get_net_stats()
            up   = max(0, (cur_sent - self.prev_sent)) // 1024
            down = max(0, (cur_recv - self.prev_recv)) // 1024
            self.prev_sent, self.prev_recv = cur_sent, cur_recv
        except Exception:
            up = down = 0
        try:    conns = get_live_connections()
        except Exception: conns = []
        try:    lan_devices = self.get_lan_now()
        except Exception: lan_devices = []
        active_lan = {}
        try:
            for dip, cnt in get_connection_bytes().items():
                if dip.startswith(self.local_subnet + ".") and dip != self.local_ip:
                    if cnt > 0:
                        self.activity_log[dip] = self.activity_log.get(dip, 0) + 2
                        active_lan[dip] = True
                    elif self.activity_log.get(dip, 0) > 0:
                        self.activity_log[dip] = max(0, self.activity_log[dip] - 2)
        except Exception:
            pass
        new_ips = {d["ip"] for d in lan_devices
                   if d["ip"] not in self.baseline_lan
                   and not d["ip"].startswith(("224.", "239.", "169.254."))}
        established = [c for c in conns if c["status"] == "ESTABLISHED"]
        listening   = [c for c in conns if c["status"] == "LISTEN"]
        other       = [c for c in conns if c["status"] not in ("ESTABLISHED", "LISTEN")]
        worst = "CRITICAL" if "CRITICAL" in self.check_status.values() else (
                "WARNING" if "WARNING" in self.check_status.values() else "OK")
        tcol  = {"OK": "bold green", "WARNING": "bold yellow", "CRITICAL": "bold red"}[worst]
        tsk   = ("tshark -> [%s]" % self.iface) if g_tshark_path else "tshark: off"
        sess  = ""
        if self.duration_end:
            tl = max(0, int(self.duration_end - time.time()))
            th, tm = divmod(tl, 3600); tm, tss = divmod(tm, 60)
            sess = "  [dim]Session ends %02d:%02d:%02d[/]" % (th, tm, tss)
        hdr = self.query_one("#hdr", Static)
        hdr.update(
            "[bold cyan]IT ANGEL[/] [bold white]WATCHMAN[/]   Next cycle [bold cyan]%02d:%02d[/]"
            "    THREAT: [%s]%s[/]    [bold green]%s[/]  [cyan]up%dKB dn%dKB/s[/]  [dim]%s[/]%s"
            % (m, s, tcol, worst, self.local_ip, up, down, tsk, sess))
        hdr.set_class(worst == "CRITICAL", "crit")
        ct = Table(box=box.SIMPLE_HEAD, expand=True, pad_edge=False, show_edge=False,
                   padding=(0, 1), title="[bold yellow]CONNECTIONS (TCP) - %d" % len(conns),
                   title_justify="left")
        ct.add_column(" ", width=1); ct.add_column("PROCESS", no_wrap=True)
        ct.add_column("SOURCE", no_wrap=True); ct.add_column("DESTINATION", no_wrap=True)
        ct.add_column("P", width=3); ct.add_column("STATUS", no_wrap=True)
        for c in (established + other + listening):
            dst_ip = c["dst"].split(":")[0] if c["dst"] else ""
            is_lan = dst_ip.startswith(self.local_subnet + ".")
            is_act = dst_ip in active_lan
            mine   = self.local_ip in c["src"]
            if is_act and is_lan:
                col = "magenta"; mk = Text("*", style="magenta")
            elif is_lan and c["status"] == "ESTABLISHED":
                col = "yellow";  mk = Text(">", style="yellow")
            elif mine and c["status"] == "ESTABLISHED":
                col = "cyan";    mk = Text(">", style="green")
            else:
                col = "white";   mk = Text(" ")
            st  = c["status"]
            stc = "green" if st == "ESTABLISHED" else ("cyan" if st == "LISTEN" else "yellow")
            ct.add_row(mk, Text(c["proc"], style=col), Text(c["src"]), Text(c["dst"]),
                       Text(c["proto"][:3]), Text(st, style=stc))
        if not conns:
            ct.add_row(" ", Text("No active connections", style="dim"), "", "", "", "")
        self.query_one("#conn", Static).update(ct)
        lt = Table(box=box.SIMPLE_HEAD, expand=True, pad_edge=False, show_edge=False,
                   padding=(0, 1),
                   title="[bold white]LAN DEVICES - %s.0/24" % self.local_subnet,
                   title_justify="left")
        lt.add_column(" ", width=1); lt.add_column("IP ADDRESS", no_wrap=True)
        lt.add_column("MAC / NOTE", no_wrap=True)
        seen = set()
        for d in lan_devices:
            ip = d["ip"]; mac = d["mac"]
            if ip.startswith(("169.254.", "224.", "239.")) or ip in seen:
                continue
            seen.add(ip)
            if ip in new_ips:
                lt.add_row(Text("*", style="bold red"), Text(ip, style="bold red"),
                           Text(mac + "  <- NEW", style="bold red"))
            elif ip in active_lan:
                lt.add_row(Text("*", style="magenta"), Text(ip, style="magenta"), Text(mac))
            elif ip == self.local_ip:
                lt.add_row(Text(">", style="green"), Text(ip, style="green"), Text(mac, style="dim"))
            else:
                lt.add_row(Text(" "), Text(ip, style="cyan"), Text(mac))
        self.query_one("#lan", Static).update(lt)
        try:
            utech, _ = check_usb_devices(self.cycle_num)
            urows = [r for r in utech if r.get("Category") == "USB Devices"]
        except Exception:
            urows = []
        usb = Table(box=box.SIMPLE_HEAD, expand=True, pad_edge=False, show_edge=False,
                    padding=(0, 1), title="[bold magenta]USB DEVICES", title_justify="left")
        usb.add_column(" ", width=1); usb.add_column("DEVICE", no_wrap=True)
        usb.add_column("STATUS", no_wrap=True)
        if not urows:
            usb.add_row(" ", Text("(no devices)", style="dim"), "")
        for r in urows:
            new = "NEW" in str(r.get("Delta", ""))
            usb.add_row(Text("*" if new else " ", style="bold red" if new else ""),
                        Text(str(r.get("Finding", ""))[:60], style="bold red" if new else "white"),
                        Text(str(r.get("Status", "")) + ("  <- NEW" if new else ""),
                             style="bold red" if new else "green"))
        self.query_one("#usb", Static).update(usb)
        self.query_one("#ftr", Static).update(
            "[white]%d established  %d listening[/]   [green]>[/] this machine  "
            "[magenta]*[/] LAN  [red]o[/] critical  [yellow]o[/] warning   "
            "[dim]| scroll up/down | q = stop | refresh 2s[/]"
            % (len(established), len(listening)))

    def action_stop(self):
        g_stop_event.set()
        self.exit()

def show_live_traffic_and_countdown(wait_seconds, cycle_num, duration_end=None,
                                    cycle_tech=None, check_status=None):
    local_ip     = get_local_ip()
    local_subnet = get_local_subnet()
    iface        = get_tshark_interface()
    start_time   = time.time()
    cycle_tech   = cycle_tech or []
    check_status = check_status or {}
    tshark_rows = []
    tshark_done = threading.Event()
    if g_tshark_path:
        threading.Thread(target=capture_tshark_background,
            args=(cycle_num, wait_seconds - 5, tshark_rows, tshark_done),
            daemon=True).start()
    _lan_cache  = list(get_lan_devices())
    _lan_lock   = threading.Lock()
    _panel_done = threading.Event()
    def _refresh_lan_loop():
        nonlocal _lan_cache
        while not _panel_done.is_set() and not g_stop_event.is_set():
            for _ in range(5):
                if _panel_done.is_set() or g_stop_event.is_set():
                    return
                time.sleep(2)
            try:
                fresh = get_lan_devices()
                with _lan_lock:
                    _lan_cache = fresh
            except Exception:
                pass
    threading.Thread(target=_refresh_lan_loop, daemon=True).start()
    def _get_lan_now():
        with _lan_lock:
            return list(_lan_cache)
    app = _WatchmanApp(wait_seconds, start_time, duration_end, local_ip, local_subnet,
                       iface, cycle_num, cycle_tech, check_status, _get_lan_now)
    try:
        app.run()
    except KeyboardInterrupt:
        g_stop_event.set()
    finally:
        _panel_done.set()
    if g_tshark_path:
        tshark_done.wait(timeout=5)
        if tshark_rows:
            print(f"  {GREEN}{len(tshark_rows)} packets -> Network Traffic sheet.{RESET}")
        else:
            print(f"  {DIM}tshark: no packets captured this cycle.{RESET}")
    print()


# --- FULL CYCLE RUNNER ---------------------------------------------------
def run_full_cycle(cycle_num):
    all_tech = []; all_exec = []
    check_status = {}
    def run_check(label, fn, *args):
        print(f"    {DIM}[ {label} ]{RESET}", end="\r")
        try:
            t, e = fn(*args)
            all_tech.extend(t); all_exec.extend(e)
            crit = any(r.get("Severity") == "CRITICAL" for r in t)
            warn = any(r.get("Severity") == "WARNING"  for r in t)
            icon = f"{RED}x{RESET}" if crit else (f"{YELLOW}!{RESET}" if warn else f"{GREEN}+{RESET}")
            stat = "CRITICAL" if crit else ("WARNING" if warn else "OK")
            check_status[label] = stat
            print(f"    {icon} {WHITE}{label:<35}{RESET} {stat}")
        except Exception as ex:
            check_status[label] = "ERR"
            print(f"    {RED}x {label}: {ex}{RESET}")
    print(f"\n  {CYAN}{'='*60}{RESET}")
    print(f"  {WHITE}Cycle {cycle_num} - {now_str()}{RESET}")
    print(f"  {CYAN}{'='*60}{RESET}\n")
    run_check("Active Connections",      check_active_connections,     cycle_num)
    run_check("Listening Ports",         check_listening_ports,        cycle_num)
    run_check("Extended Port Scan",      check_extended_ports,         cycle_num)
    run_check("Net Processes",           check_network_processes,      cycle_num)
    run_check("Privileged Processes",    check_privileged_processes,   cycle_num)
    run_check("Resources (CPU/RAM)",     check_resources,              cycle_num)
    run_check("Logged Users",            check_logged_users,           cycle_num)
    run_check("Local Users & Admins",    check_local_users,            cycle_num)
    run_check("RDP / Remote Sessions",   check_rdp_sessions,           cycle_num)
    run_check("Services",                check_services,               cycle_num)
    run_check("Firewall",                check_firewall,               cycle_num)
    run_check("USB Devices",             check_usb_devices,            cycle_num)
    run_check("DNS & Gateway",           check_dns_gateway,            cycle_num)
    run_check("Scheduled Tasks",         check_scheduled_tasks,        cycle_num)
    run_check("LAN Devices (ARP)",       check_local_network,          cycle_num)
    run_check("System Events",           check_system_events,          cycle_num)
    run_check("Security Events Detail",  check_security_events_detail, cycle_num)
    run_check("Network Shares",          check_open_shares,            cycle_num)
    run_check("Startup Items",           check_startup_items,          cycle_num)
    run_check("Recent Software",         check_recent_software,        cycle_num)
    run_check("File Integrity",          check_system_file_integrity,  cycle_num)
    run_check("Defender / AV",           check_defender_status,        cycle_num)
    run_check("Windows Update",          check_windows_update,         cycle_num)
    run_check("BitLocker / SecureBoot",  check_bitlocker_secureboot,   cycle_num)
    run_check("Audit Policy",            check_audit_policy,           cycle_num)
    run_check("Event Log Cleared",       check_event_log_cleared,      cycle_num)
    run_check("Shadow Copies",           check_shadow_copies,          cycle_num)
    run_check("System Time",             check_system_time,            cycle_num)
    run_check("Registry Autoruns",       check_registry_autoruns,      cycle_num)
    run_check("WMI Subscriptions",       check_wmi_subscriptions,      cycle_num)
    run_check("Browser Extensions",      check_browser_extensions,     cycle_num)
    run_check("Beaconing",               check_beaconing,              cycle_num)
    run_check("ARP / Gateway",           check_arp_spoofing,           cycle_num)
    run_check("Process Lineage",         check_process_lineage,        cycle_num)
    run_check("Ransomware IOC",          check_ransomware_indicators,  cycle_num)
    run_check("Webcam / Mic",            check_webcam_mic,             cycle_num)
    print(f"\n  {DIM}Saving to Excel...{RESET}", end="\r")
    append_to_excel(all_tech, all_exec)
    print(f"  {GREEN}Excel updated - {g_excel_path}{RESET}")
    return all_tech, all_exec, check_status


# --- BASELINE ------------------------------------------------------------
def run_baseline():
    print(f"\n  {CYAN}Running initial system baseline...{RESET}\n")
    all_tech = []; all_exec = []
    for label, fn in [("System Snapshot",  check_system_snapshot),
                      ("Network Adapters", check_network_adapters)]:
        print(f"    {DIM}[ {label} ]{RESET}", end="\r")
        try:
            t, e = fn()
            all_tech.extend(t); all_exec.extend(e)
            print(f"    {GREEN}+{RESET} {WHITE}{label}{RESET}")
        except Exception as ex:
            print(f"    {RED}x {label}: {ex}{RESET}")
    append_to_excel(all_tech, all_exec)
    silent = [
        ("Ports baseline",     check_listening_ports),
        ("Processes baseline", check_network_processes),
        ("Services baseline",  check_services),
        ("Tasks baseline",     check_scheduled_tasks),
        ("USB baseline",       check_usb_devices),
    ]
    for label, fn in silent:
        print(f"    {DIM}[ {label} ]{RESET}", end="\r")
        try:
            fn("BASELINE")
            print(f"    {GREEN}+{RESET} {WHITE}{label}{RESET}")
        except Exception as ex:
            print(f"    {RED}x {label}: {ex}{RESET}")
    try: check_local_network("BASELINE")
    except: pass
    print(f"\n  {GREEN}Baseline complete - Cycle 1 will be the first active check.{RESET}\n")


# --- FAST CHECK MODE -----------------------------------------------------
def run_fast_check():
    print(f"\n  {MAGENTA}Fast Check - running all checks once...{RESET}\n")
    all_tech = []; all_exec = []
    for label, fn in [("System Snapshot", check_system_snapshot),
                      ("Network Adapters", check_network_adapters)]:
        try:
            t, e = fn()
            all_tech.extend(t); all_exec.extend(e)
            print(f"    {GREEN}+{RESET} {WHITE}{label}{RESET}")
        except: pass
    append_to_excel(all_tech, all_exec)
    run_full_cycle("FAST")
    if g_tshark_path:
        print(f"\n  {CYAN}Capturing 15s network sample with tshark...{RESET}")
        fast_rows = []
        fast_done = threading.Event()
        t = threading.Thread(target=capture_tshark_background,
                             args=("FAST", 15, fast_rows, fast_done), daemon=True)
        t.start()
        fast_done.wait(timeout=30)
        if fast_rows:
            save_traffic_to_excel(fast_rows)
            print(f"  {GREEN}{len(fast_rows)} packets captured - Network Traffic sheet.{RESET}")
    print(f"\n  {GREEN}{'='*60}{RESET}")
    print(f"  {MAGENTA}YOUR REPORT IS READY!{RESET}")
    print(f"  {GREEN}  {g_excel_path}{RESET}")
    print(f"  {GREEN}{'='*60}{RESET}\n")


# --- MONITORING LOOP -----------------------------------------------------
def monitoring_loop():
    start_time   = time.time()
    duration_end = (start_time + g_duration) if g_duration > 0 else None
    cycle_num    = 1
    print(f"\n  {WHITE}{'='*60}{RESET}")
    print(f"  {GREEN} IT Angel is now protecting this system. Ctrl+C to stop.{RESET}")
    print(f"  {WHITE}{'='*60}{RESET}")
    try:
        while not g_stop_event.is_set():
            if duration_end and time.time() >= duration_end:
                print(f"\n  {YELLOW}Protection period complete.{RESET}")
                break
            _all_tech, _all_exec, _check_status = run_full_cycle(cycle_num)
            cycle_num += 1
            if duration_end and time.time() >= duration_end:
                break
            show_live_traffic_and_countdown(CYCLE_INTERVAL, cycle_num-1, duration_end,
                                            _all_tech, _check_status)
    except KeyboardInterrupt:
        pass
    print(f"\n  {GREEN}{'='*60}{RESET}")
    print(f"  {WHITE} Session complete. {cycle_num-1} cycle(s) executed.{RESET}")
    print(f"  {GREEN} Report saved to:{RESET}")
    print(f"  {CYAN}  {g_excel_path}{RESET}")
    print(f"  {GREEN}{'='*60}{RESET}\n")


# --- MAIN MENU -----------------------------------------------------------
def main():
    while True:
        print_banner()
        if detect_tshark():
            print(f"  {GREEN}tshark detected:{RESET} {g_tshark_path}\n")
        else:
            print(f"  {YELLOW}tshark not found. Live traffic monitoring will be disabled.")
            print(f"    Install Wireshark (Windows) or: sudo apt install tshark (Linux){RESET}\n")
        select_os()
        select_duration()
        if g_duration == -1:
            setup_excel()
            run_fast_check()
            print(f"  {WHITE}Press Enter to run another check or Ctrl+C to exit.{RESET}")
            try:
                input()
            except (KeyboardInterrupt, EOFError):
                break
            continue
        setup_excel()
        run_baseline()
        monitoring_loop()
        print(f"  {WHITE}Press Enter to start a new session or Ctrl+C to exit.{RESET}")
        try:
            input()
        except (KeyboardInterrupt, EOFError):
            break


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print(f"\n\n  {YELLOW}IT Angel stopped. Stay protected.{RESET}\n")
